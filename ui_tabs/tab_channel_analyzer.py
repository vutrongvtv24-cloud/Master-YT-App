from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QTextEdit, QTableWidget,
    QTableWidgetItem, QHeaderView, QFileDialog, QMessageBox
)
from PyQt6.QtCore import Qt, QRunnable, pyqtSignal, QObject, QUrl, QThreadPool
from PyQt6.QtGui import QDesktopServices, QColor
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
import pandas as pd
from datetime import datetime
import logging
import re
import time
from openpyxl import load_workbook
from utils import extract_channel_id_yt_dlp, format_number, format_date_dd_mm_yyyy

logger = logging.getLogger(__name__)

class ChannelAnalyzerSignals(QObject):
    data_fetched = pyqtSignal(list, bool)  # list of channel data, success flag
    status_updated = pyqtSignal(str, int)  # message, timeout
    error_occurred = pyqtSignal(str)
    progress_updated = pyqtSignal(int, str)  # value, message

class ChannelAnalyzerRunnable(QRunnable):
    def __init__(self, api_key, channel_urls, parent):
        super().__init__()
        self.signals = ChannelAnalyzerSignals()
        self.api_key = api_key
        self.channel_urls = channel_urls
        self.parent = parent
        self._is_interruption_requested = False
        self.setAutoDelete(True)  # Automatically delete runnable after execution

    def is_valid_youtube_url(self, url):
        """Kiểm tra định dạng URL YouTube hợp lệ."""
        pattern = r'^(https?:\/\/)?(www\.)?(youtube\.com|youtu\.be)\/(channel\/|user\/|@|c\/)?[\w\-]+'
        return bool(re.match(pattern, url))

    def run(self):
        if not self.api_key:
            self.signals.status_updated.emit("API Key không có hoặc không hợp lệ.", 3000)
            self.signals.data_fetched.emit([], False)
            return

        self.signals.progress_updated.emit(0, "Đang khởi tạo...")
        logger.debug("Starting channel analysis with %d URLs", len(self.channel_urls))
        results = []
        total_urls = len(self.channel_urls)
        processed = 0

        try:
            from youtube_service import YouTubeService, APIKeyManager
            key_list = self.api_key.split('\n')
            key_manager = APIKeyManager(key_list)
            youtube_service_wrapper = YouTubeService(key_manager)
            channel_ids = []
            url_to_info = {}

            # Step 1: Extract channel IDs
            self.signals.progress_updated.emit(5, "Đang trích xuất ID kênh...")
            for url in self.channel_urls:
                if self.isInterruptionRequested():
                    self.signals.status_updated.emit("Hủy phân tích kênh.", 2000)
                    self.signals.data_fetched.emit(results, False)
                    return
                if not self.is_valid_youtube_url(url):
                    results.append({
                        'name': 'N/A',
                        'subscribers': 'N/A',
                        'video_count': 'N/A',
                        'view_count': 'N/A',
                        'created_date': 'N/A',
                        'country': 'N/A',
                        'category': 'N/A',
                        'url': url,
                        'status': 'URL không hợp lệ'
                    })
                    processed += 1
                    progress = int(5 + (processed / total_urls) * 45)
                    self.signals.progress_updated.emit(progress, f"Đã xử lý {processed}/{total_urls} URL...")
                    continue
                channel_id, error = extract_channel_id_yt_dlp(url, lambda msg, timeout: self.signals.status_updated.emit(msg, timeout))
                if channel_id:
                    channel_ids.append(channel_id)
                    url_to_info[channel_id] = {'url': url}
                else:
                    results.append({
                        'name': 'N/A',
                        'subscribers': 'N/A',
                        'video_count': 'N/A',
                        'view_count': 'N/A',
                        'created_date': 'N/A',
                        'country': 'N/A',
                        'category': 'N/A',
                        'url': url,
                        'status': error or 'Không trích xuất được ID kênh'
                    })
                processed += 1
                progress = int(5 + (processed / total_urls) * 45)
                self.signals.progress_updated.emit(progress, f"Đã xử lý {processed}/{total_urls} URL...")
                logger.debug("Processed URL %s, channel_id: %s", url, channel_id)

            if not channel_ids:
                self.signals.status_updated.emit("Không tìm thấy ID kênh hợp lệ.", 3000)
                self.signals.data_fetched.emit(results, False)
                return

            # Step 2: Fetch channel data in batches
            batch_size = 50
            for i in range(0, len(channel_ids), batch_size):
                if self.isInterruptionRequested():
                    self.signals.status_updated.emit("Hủy phân tích kênh.", 2000)
                    self.signals.data_fetched.emit(results, False)
                    return
                batch_ids = channel_ids[i:i + batch_size]
                self.signals.progress_updated.emit(50 + (i / len(channel_ids)) * 40, f"Đang lấy dữ liệu cho {len(batch_ids)} kênh...")
                
                # Wrapper method handles retry and rotation automatically
                try:
                    response = youtube_service_wrapper.get_channel_details(batch_ids)
                    for item in response.get('items', []):
                        channel_id = item['id']
                        snippet = item.get('snippet', {})
                        stats = item.get('statistics', {})
                        topic_details = item.get('topicDetails', {})
                        subscribers = int(stats.get('subscriberCount', '0')) if stats.get('subscriberCount', '0').isdigit() else 0
                        video_count = int(stats.get('videoCount', '0')) if stats.get('videoCount', '0').isdigit() else 0
                        view_count = int(stats.get('viewCount', '0')) if stats.get('viewCount', '0').isdigit() else 0
                        categories = topic_details.get('topicCategories', []) or ['N/A']
                        categories = [cat.split('/')[-1] for cat in categories]
                        results.append({
                            'name': snippet.get('title', 'N/A'),
                            'subscribers': subscribers,
                            'video_count': video_count,
                            'view_count': view_count,
                            'created_date': format_date_dd_mm_yyyy(snippet.get('publishedAt', 'N/A')),
                            'country': snippet.get('country', 'N/A'),
                            'category': ', '.join(categories),
                            'url': url_to_info[channel_id]['url'],
                            'status': 'Thành công'
                        })
                except HttpError as e:
                    error_msg = f"Lỗi API: {str(e)}"
                    if e.resp.status in (403, 400):
                        error_msg = "Hết quota hoặc API Key không hợp lệ."
                        for channel_id in batch_ids:
                            results.append({
                                'name': 'N/A',
                                'subscribers': 0,
                                'video_count': 0,
                                'view_count': 0,
                                'created_date': 'N/A',
                                'country': 'N/A',
                                'category': 'N/A',
                                'url': url_to_info.get(channel_id, {}).get('url', 'N/A'),
                                'status': error_msg
                            })
                        self.signals.error_occurred.emit(error_msg)
                    else:
                        # General error handling
                        for channel_id in batch_ids:
                           results.append({
                               'name': 'N/A',
                               'subscribers': 0,
                               'video_count': 0,
                               'view_count': 0,
                               'created_date': 'N/A',
                               'country': 'N/A',
                               'category': 'N/A',
                               'url': url_to_info.get(channel_id, {}).get('url', 'N/A'),
                               'status': error_msg
                           })
                        self.signals.error_occurred.emit(error_msg)



                processed += len(batch_ids)
                progress = int(50 + (processed / total_urls) * 40)
                self.signals.progress_updated.emit(progress, f"Đã lấy dữ liệu {processed}/{total_urls} kênh...")
                logger.debug("Processed batch %d/%d", i + batch_size, len(channel_ids))

            self.signals.progress_updated.emit(100, "Hoàn tất phân tích kênh.")
            success_count = sum(1 for r in results if r['status'] == 'Thành công')
            self.signals.status_updated.emit(
                f"Đã phân tích {success_count}/{total_urls} kênh thành công, {total_urls - success_count} lỗi.", 5000
            )
            self.signals.data_fetched.emit(results, True)

        except Exception as e:
            logger.error(f"Lỗi nghiêm trọng: {str(e)}", exc_info=True)
            self.signals.error_occurred.emit(f"Lỗi nghiêm trọng: {str(e)}")
            self.signals.progress_updated.emit(100, f"Thất bại: {str(e)}")
            self.signals.data_fetched.emit(results, False)

    def requestInterruption(self):
        self._is_interruption_requested = True

    def isInterruptionRequested(self):
        return self._is_interruption_requested

class ChannelAnalyzerTab(QWidget):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        self.results = []
        self.runnable = None
        self.thread_pool = QThreadPool.globalInstance()
        self.init_ui()
        self.apply_styles()

    def init_ui(self):
        layout = QVBoxLayout()

        # URL input
        self.url_input = QTextEdit()
        self.url_input.setPlaceholderText("Nhập danh sách URL kênh, mỗi URL trên một dòng...")
        self.url_input.setFixedHeight(100)
        self.url_input.textChanged.connect(self.update_url_count)
        layout.addWidget(self.url_input)

        # URL count label
        self.url_count_label = QTextEdit()
        self.url_count_label.setReadOnly(True)
        self.url_count_label.setFixedHeight(30)
        self.url_count_label.setText("Số URL: 0")
        layout.addWidget(self.url_count_label)

        # Buttons
        button_layout = QHBoxLayout()
        self.analyze_button = QPushButton("Phân tích Kênh")
        self.analyze_button.clicked.connect(self.start_analysis)
        button_layout.addWidget(self.analyze_button)

        self.cancel_button = QPushButton("Hủy")
        self.cancel_button.clicked.connect(self.cancel_analysis)
        self.cancel_button.setEnabled(False)
        button_layout.addWidget(self.cancel_button)

        self.clear_button = QPushButton("Xóa danh sách URL")
        self.clear_button.clicked.connect(self.clear_urls)
        button_layout.addWidget(self.clear_button)

        self.export_button = QPushButton("Xuất Excel")
        self.export_button.clicked.connect(self.export_to_excel)
        self.export_button.setEnabled(False)
        button_layout.addWidget(self.export_button)

        layout.addLayout(button_layout)

        # Results table
        self.table = QTableWidget()
        self.table.setColumnCount(9)
        self.table.setHorizontalHeaderLabels([
            "Tên kênh", "Người theo dõi", "Số Video", "Tổng Lượt Xem",
            "Ngày tạo Kênh", "Quốc gia", "Danh mục", "URL kênh", "Trạng thái"
        ])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setSortingEnabled(True)
        self.table.cellClicked.connect(self.handle_cell_click)
        layout.addWidget(self.table)

        self.setLayout(layout)

    def apply_styles(self):
        self.analyze_button.setObjectName("btnPrimary")
        self.cancel_button.setObjectName("btnSecondary")
        self.clear_button.setObjectName("btnSecondary")
        self.export_button.setObjectName("btnSecondary")
        self.table.setStyleSheet("""
            QTableWidget::item { padding: 5px; }
            QTableWidget::item:selected { background-color: #0078D7; color: white; }
        """)
        self.url_count_label.setStyleSheet("QTextEdit { background-color: #F0F0F0; border: 1px solid #CCCCCC; }")

    def update_url_count(self):
        urls = [url.strip() for url in self.url_input.toPlainText().split('\n') if url.strip()]
        self.url_count_label.setText(f"Số URL: {len(urls)}")

    def clear_urls(self):
        self.url_input.clear()
        self.parent.statusBar().showMessage("Đã xóa danh sách URL.", 3000)

    def cancel_analysis(self):
        if self.runnable:
            self.runnable.requestInterruption()
            self.cancel_button.setEnabled(False)
            self.analyze_button.setEnabled(True)
            self.parent.statusBar().showMessage("Đang hủy phân tích...", 3000)

    def start_analysis(self):
        if self.parent.is_operation_running:
            self.parent.statusBar().showMessage("Một tác vụ khác đang chạy, vui lòng đợi.", 3000)
            return

        urls = [url.strip() for url in self.url_input.toPlainText().split('\n') if url.strip()]
        if not urls:
            self.parent.statusBar().showMessage("Vui lòng nhập ít nhất một URL kênh.", 3000)
            return

        self.parent.set_operation_running_status(True, "Phân tích kênh")
        self.parent.show_progress_dialog("Đang khởi tạo phân tích kênh...")
        self.table.setRowCount(0)
        self.results = []
        self.analyze_button.setEnabled(False)
        self.cancel_button.setEnabled(True)

        # Process in batches of 50 URLs
        self.current_batch = 0
        self.url_batches = [urls[i:i+50] for i in range(0, len(urls), 50)]
        self.run_next_batch()

    def run_next_batch(self):
        if self.current_batch >= len(self.url_batches):
            self.on_all_batches_finished()
            return

        batch_urls = self.url_batches[self.current_batch]
        self.runnable = ChannelAnalyzerRunnable(self.parent.get_active_api_key(), batch_urls, self.parent)
        self.runnable.signals.data_fetched.connect(self.on_batch_data_fetched)
        self.runnable.signals.status_updated.connect(lambda msg, timeout: self.parent.statusBar().showMessage(msg, timeout))
        self.runnable.signals.error_occurred.connect(self.parent.on_api_error_common_slot)
        self.runnable.signals.progress_updated.connect(self.parent.update_progress_dialog)
        self.parent.worker_started(self.runnable, "Phân tích kênh")
        self.thread_pool.start(self.runnable)  # Explicitly start the runnable
        self.current_batch += 1

    def on_batch_data_fetched(self, batch_results, success):
        self.results.extend(batch_results)
        self.update_table()
        self.run_next_batch()

    def on_all_batches_finished(self):
        self.export_button.setEnabled(bool(self.results))
        self.analyze_button.setEnabled(True)
        self.cancel_button.setEnabled(False)
        self.parent.on_worker_thread_finished()

    def update_table(self):
        self.table.setRowCount(len(self.results))
        for row, data in enumerate(self.results):
            self.table.setItem(row, 0, QTableWidgetItem(data['name']))
            self.table.setItem(row, 1, QTableWidgetItem(format_number(data['subscribers'])))
            self.table.setItem(row, 2, QTableWidgetItem(format_number(data['video_count'])))
            self.table.setItem(row, 3, QTableWidgetItem(format_number(data['view_count'])))
            self.table.setItem(row, 4, QTableWidgetItem(data['created_date']))
            self.table.setItem(row, 5, QTableWidgetItem(data['country']))
            self.table.setItem(row, 6, QTableWidgetItem(data['category']))
            url_item = QTableWidgetItem(data['url'])
            url_item.setForeground(QColor(0, 0, 255))
            url_item.setData(Qt.ItemDataRole.UserRole, data['url'])
            self.table.setItem(row, 7, url_item)
            self.table.setItem(row, 8, QTableWidgetItem(data['status']))

    def handle_cell_click(self, row, column):
        if column == 7:  # URL column
            item = self.table.item(row, column)
            if item:
                url = item.data(Qt.ItemDataRole.UserRole)
                if url:
                    self.parent.open_url_externally(url)

    def export_to_excel(self):
        if not self.results:
            self.parent.statusBar().showMessage("Không có dữ liệu để xuất.", 3000)
            return

        current_date = datetime.now().strftime("%d-%m")
        default_filename = f"PT Kenh {current_date}.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Lưu file Excel", default_filename, "Excel Files (*.xlsx)"
        )
        if not file_path:
            return

        try:
            # Check if file is writable
            with open(file_path, 'a'):
                pass
            # Create DataFrame with desired column order
            df = pd.DataFrame(self.results, columns=[
                'name', 'subscribers', 'video_count', 'view_count', 'created_date',
                'country', 'category', 'url', 'status'
            ])
            df.to_excel(file_path, index=False, engine='openpyxl')

            # Add hyperlinks to URL column
            wb = load_workbook(file_path)
            ws = wb.active
            for row in range(2, len(self.results) + 2):
                url = self.results[row - 2]['url']
                if url != 'N/A':
                    ws.cell(row=row, column=8, value=url)
                    ws.cell(row=row, column=8).hyperlink = url
                    ws.cell(row=row, column=8).style = 'Hyperlink'

            wb.save(file_path)
            self.parent.statusBar().showMessage(f"Đã xuất file Excel: {file_path}", 5000)
        except PermissionError:
            self.parent.statusBar().showMessage("Lỗi: File Excel đang mở, vui lòng đóng trước khi xuất.", 5000)
        except Exception as e:
            self.parent.statusBar().showMessage(f"Lỗi khi xuất Excel: {str(e)}", 5000)
            logger.error(f"Lỗi xuất Excel: {str(e)}", exc_info=True)

    def set_buttons_enabled(self, enabled):
        self.analyze_button.setEnabled(enabled)
        self.cancel_button.setEnabled(False)
        self.clear_button.setEnabled(enabled)
        self.export_button.setEnabled(enabled and bool(self.results))