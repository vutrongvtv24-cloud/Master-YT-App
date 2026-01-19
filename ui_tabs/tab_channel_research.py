# youtube_research_tool/ui_tabs/tab_channel_research.py

import json
import traceback
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QTextEdit,
    QPushButton, QTableWidget, QTableWidgetItem, QMessageBox,
    QFileDialog, QGroupBox, QHeaderView, QMenu, QApplication,
    QLineEdit
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QUrl
from PyQt6.QtGui import QDesktopServices, QClipboard, QFont, QColor, QIntValidator

from utils import (
    extract_channel_id_yt_dlp,
    format_datetime_iso,
    format_int_with_separator,
    convert_iso_duration
)

from googleapiclient.discovery import build
from googleapiclient.errors import HttpError

def format_date_to_ddmmyyyy(iso_date):
    if not iso_date:
        return 'N/A'
    try:
        return datetime.fromisoformat(iso_date[:-1]).strftime('%d-%m-%Y')
    except ValueError:
        return 'N/A'

def _parse_duration_to_seconds(duration_str):
    """Hàm phụ trợ để chuyển đổi chuỗi thời lượng (HH:MM:SS, MM:SS, SS) thành giây."""
    if not duration_str or not isinstance(duration_str, str):
        return 0
    parts = duration_str.strip().split(':')
    seconds = 0
    try:
        if len(parts) == 1:
            seconds = int(parts[0])
        elif len(parts) == 2:
            seconds = int(parts[0]) * 60 + int(parts[1])
        elif len(parts) == 3:
            seconds = int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])
    except (ValueError, IndexError):
        return 0
    return seconds


class FetchChannelVideosThread(QThread):
    channel_videos_fetched = pyqtSignal(list, str)
    error_occurred = pyqtSignal(str)
    progress_updated = pyqtSignal(int, str)

    def __init__(self, api_key, channel_id, video_categories_map, parent=None):
        super().__init__(parent)
        self.api_key = api_key
        self.channel_id = channel_id
        self.video_categories_map = video_categories_map or {}
        self._is_interruption_requested = False

    def run(self):
        if not self.api_key:
            self.error_occurred.emit("Vui lòng cung cấp API Key ở Tab 1 (API Key).")
            return
        if not self.channel_id:
            self.error_occurred.emit("Channel ID không hợp lệ.")
            return

        try:
            self.progress_updated.emit(0, f"Đang kết nối tới YouTube để lấy thông tin kênh {self.channel_id}...")
            self.progress_updated.emit(0, f"Đang kết nối tới YouTube để lấy thông tin kênh {self.channel_id}...")
            from youtube_service import YouTubeService, APIKeyManager

            key_list = self.api_key.split('\n')
            key_manager = APIKeyManager(key_list)
            youtube_service_wrapper = YouTubeService(key_manager)
            
            if self.isInterruptionRequested(): return

            self.progress_updated.emit(5, f"Đang lấy thông tin chi tiết kênh {self.channel_id}...")
            channel_response = youtube_service_wrapper.get_channel_details(self.channel_id)
            if self.isInterruptionRequested(): return

            if not channel_response.get("items"):
                self.error_occurred.emit(f"Không tìm thấy kênh với ID: {self.channel_id}")
                return
            
            channel_item = channel_response["items"][0]
            channel_title = channel_item.get("snippet", {}).get("title", f"Kênh {self.channel_id}")
            uploads_playlist_id = channel_item.get("contentDetails", {}).get("relatedPlaylists", {}).get("uploads")

            if not uploads_playlist_id:
                self.error_occurred.emit(f"Không tìm thấy playlist video tải lên cho kênh: {channel_title} ({self.channel_id})")
                return
            
            self.progress_updated.emit(10, f"Đã tìm thấy playlist tải lên cho kênh '{channel_title}'. Đang lấy danh sách ID video...")

            all_video_item_details_from_playlist = []
            next_page_token = None
            item_count_from_playlist = 0
            while True:
                if self.isInterruptionRequested(): return
                playlist_response = youtube_service_wrapper.get_playlist_items(
                    playlist_id=uploads_playlist_id,
                    max_results=50,
                    page_token=next_page_token
                )
                if self.isInterruptionRequested(): return

                current_items = playlist_response.get("items", [])
                for item in current_items:
                    if item.get("snippet", {}).get("resourceId", {}).get("kind") == "youtube#video" and \
                       item.get("snippet", {}).get("resourceId", {}).get("videoId"):
                        all_video_item_details_from_playlist.append(item)
                
                item_count_from_playlist += len(current_items)
                current_playlist_progress = 10 + min(30, item_count_from_playlist // 10 if item_count_from_playlist > 0 else 0)
                self.progress_updated.emit(current_playlist_progress, f"Đang lấy ID video từ kênh ({len(all_video_item_details_from_playlist)} đã tìm thấy)...")

                next_page_token = playlist_response.get("nextPageToken")
                if not next_page_token or not current_items:
                    break
            
            if not all_video_item_details_from_playlist:
                self.channel_videos_fetched.emit([], channel_title)
                self.progress_updated.emit(100, f"Kênh '{channel_title}' không có video nào (hoặc không thể truy cập).")
                return

            if self.isInterruptionRequested(): return
            self.progress_updated.emit(40, f"Đã lấy {len(all_video_item_details_from_playlist)} ID video. Đang lấy chi tiết thống kê và danh mục...")

            final_video_data = []
            video_ids_to_fetch_details = [item["snippet"]["resourceId"]["videoId"] for item in all_video_item_details_from_playlist]

            for i in range(0, len(video_ids_to_fetch_details), 50):
                if self.isInterruptionRequested(): return
                chunk_video_ids = video_ids_to_fetch_details[i:i+50]
                
                videos_list_response = youtube_service_wrapper.get_video_details(chunk_video_ids)
                if self.isInterruptionRequested(): return

                for video_stat_item in videos_list_response.get("items", []):
                    video_id = video_stat_item["id"]
                    snippet = video_stat_item.get("snippet", {})
                    statistics = video_stat_item.get("statistics", {})
                    content_details = video_stat_item.get("contentDetails", {})
                    
                    category_id = snippet.get('categoryId', '')
                    category_name = 'Không xác định'
                    if self.video_categories_map:
                        category_name = next((name for name, cid_val in self.video_categories_map.items() if cid_val == category_id), 'Không xác định')
                    
                    final_video_data.append({
                        'id': video_id,
                        'title': snippet.get('title', 'N/A'),
                        'url': f"https://www.youtube.com/watch?v={video_id}",
                        'view_count': int(statistics.get('viewCount', 0)),
                        'comment_count': statistics.get('commentCount'),
                        'upload_date': snippet.get('publishedAt', ''),
                        'duration': convert_iso_duration(content_details.get('duration', 'N/A')),
                        'category_name': category_name
                    })
                
                details_progress = 40 + int(55 * (len(final_video_data) / len(all_video_item_details_from_playlist)))
                self.progress_updated.emit(details_progress, f"Đang lấy chi tiết video cho kênh ({len(final_video_data)}/{len(all_video_item_details_from_playlist)})...")

            if self.isInterruptionRequested(): return
            self.progress_updated.emit(100, f"Hoàn tất lấy video cho kênh '{channel_title}'.")
            self.channel_videos_fetched.emit(final_video_data, channel_title)

        except HttpError as e:
            if self.isInterruptionRequested(): return
            try:
                error_content = json.loads(e.content.decode('utf-8'))
                error_message = error_content.get("error", {}).get("message", "Lỗi API không xác định khi lấy video kênh.")
                status_code = e.resp.status
                if status_code == 403 and ("quotaExceeded" in error_message or "dailyLimitExceeded" in error_message):
                    self.error_occurred.emit(f"Lỗi hạn ngạch API khi xử lý kênh {self.channel_id}.")
                elif status_code == 404:
                     self.error_occurred.emit(f"Không tìm thấy kênh hoặc playlist với ID cung cấp ({self.channel_id}).")
                else:
                    self.error_occurred.emit(f"Lỗi API ({status_code}) khi lấy video kênh: {error_message}")
            except json.JSONDecodeError:
                self.error_occurred.emit(f"Lỗi API (không thể phân tích phản hồi) khi lấy video kênh: {e.content.decode('utf-8', errors='ignore')}")
        except Exception as e:
            if self.isInterruptionRequested(): return
            traceback.print_exc()
            self.error_occurred.emit(f"Lỗi không mong đợi khi lấy video kênh '{self.channel_id}': {str(e)}. Xem console.")

    def requestInterruption(self):
        self._is_interruption_requested = True
        super().requestInterruption()

    def isInterruptionRequested(self):
        return self._is_interruption_requested or super().isInterruptionRequested()

class ChannelResearchTab(QWidget):
    def __init__(self, main_window):
        super().__init__()
        self.main_window = main_window
        self.fetch_channel_videos_thread = None
        self.channel_videos_data = {}  # Lưu trữ toàn bộ dữ liệu video chưa lọc
        self.current_channel_names_for_export = []

        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        # Nhóm nhập liệu
        input_group = QGroupBox("Nhập liệu Phân tích Kênh")
        input_group.setFont(QFont("Arial", 11, QFont.Weight.Bold))
        input_group.setStyleSheet("QGroupBox { background-color: #F5F5F5; }")
        input_layout = QVBoxLayout()
        input_layout.setSpacing(10)

        url_layout = QHBoxLayout()
        url_label = QLabel("URL hoặc ID Kênh YouTube (mỗi kênh một dòng):")
        url_label.setFont(QFont("Arial", 10))
        url_layout.addWidget(url_label)
        self.txt_channel_urls = QTextEdit()
        self.txt_channel_urls.setFont(QFont("Arial", 10))
        self.txt_channel_urls.setPlaceholderText("Dán URL kênh (ví dụ: @MrBeast, youtube.com/channel/ID) hoặc ID kênh (UC...)\nMỗi kênh một dòng.")
        self.txt_channel_urls.setAcceptRichText(False)
        self.txt_channel_urls.setMinimumHeight(120)
        self.txt_channel_urls.setStyleSheet("QTextEdit { padding: 5px; }")
        url_layout.addWidget(self.txt_channel_urls, 1)
        input_layout.addLayout(url_layout)

        self.lbl_channel_count = QLabel("Số kênh: 0")
        self.lbl_channel_count.setFont(QFont("Arial", 9))
        self.txt_channel_urls.textChanged.connect(self._update_channel_count)
        input_layout.addWidget(self.lbl_channel_count)

        analyze_button_layout = QHBoxLayout()
        analyze_button_layout.addStretch()
        self.btn_analyze_channel = QPushButton("Phân tích Kênh")
        self.btn_analyze_channel.setFont(QFont("Arial", 10))
        self.btn_analyze_channel.setToolTip("Phân tích tất cả các kênh được nhập để lấy danh sách video")
        self.btn_analyze_channel.clicked.connect(self._start_analyze_channel_videos)
        self.btn_analyze_channel.setFixedWidth(150)
        analyze_button_layout.addWidget(self.btn_analyze_channel)
        input_layout.addLayout(analyze_button_layout)
        input_group.setLayout(input_layout)
        layout.addWidget(input_group)

        # --- BỘ LỌC ---
        filter_group = QGroupBox("Bộ lọc")
        filter_group.setFont(QFont("Arial", 11, QFont.Weight.Bold))
        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(15)

        # Lượt xem tối thiểu
        filter_layout.addWidget(QLabel("Lượt xem ≥"))
        self.le_min_views = QLineEdit()
        self.le_min_views.setValidator(QIntValidator(0, 2000000000))
        self.le_min_views.setPlaceholderText("VD: 10000")
        self.le_min_views.setFixedWidth(120)
        filter_layout.addWidget(self.le_min_views)

        # Bình luận tối thiểu
        filter_layout.addWidget(QLabel("Bình luận ≥"))
        self.le_min_comments = QLineEdit()
        self.le_min_comments.setValidator(QIntValidator(0, 10000000))
        self.le_min_comments.setPlaceholderText("VD: 100")
        self.le_min_comments.setFixedWidth(120)
        filter_layout.addWidget(self.le_min_comments)

        # Thời lượng tối thiểu
        filter_layout.addWidget(QLabel("Thời lượng (phút) ≥"))
        self.le_min_duration_minutes = QLineEdit()
        self.le_min_duration_minutes.setValidator(QIntValidator(0, 999999))
        self.le_min_duration_minutes.setPlaceholderText("VD: 5")
        self.le_min_duration_minutes.setFixedWidth(120)
        filter_layout.addWidget(self.le_min_duration_minutes)

        filter_layout.addStretch()

        self.btn_apply_filters = QPushButton("Áp dụng Bộ lọc")
        self.btn_apply_filters.setFont(QFont("Arial", 10))
        self.btn_apply_filters.setToolTip("Áp dụng các bộ lọc trên vào danh sách video đã lấy")
        self.btn_apply_filters.setFixedWidth(150)
        self.btn_apply_filters.clicked.connect(self._update_display_with_filters)
        filter_layout.addWidget(self.btn_apply_filters)

        filter_group.setLayout(filter_layout)
        layout.addWidget(filter_group)
        # --- KẾT THÚC BỘ LỌC ---

        # Nhóm kết quả
        results_group = QGroupBox("Danh sách Video của Kênh")
        results_group.setFont(QFont("Arial", 11, QFont.Weight.Bold))
        results_group.setStyleSheet("QGroupBox { background-color: #F5F5F5; }")
        results_layout = QVBoxLayout()
        results_layout.setSpacing(10)

        self.table_channel_videos = QTableWidget()
        self.table_channel_videos.setColumnCount(9)
        self.table_channel_videos.setHorizontalHeaderLabels([
            "Tên Kênh", "Tiêu đề", "Lượt xem", "Bình luận", "Ngày đăng", "Thời lượng", 
            "Danh mục", "URL Video", "Hành động"
        ])
        self.table_channel_videos.setFont(QFont("Arial", 10))
        self.table_channel_videos.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table_channel_videos.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table_channel_videos.setSortingEnabled(True)
        
        header_channel = self.table_channel_videos.horizontalHeader()
        header_channel.setFont(QFont("Arial", 10, QFont.Weight.Bold))
        col_widths = [180, -1, 120, 120, 120, 120, 150, 180, 80]
        for i, width in enumerate(col_widths):
            if width == -1:
                header_channel.setSectionResizeMode(i, QHeaderView.ResizeMode.Stretch)
            else:
                header_channel.setSectionResizeMode(i, QHeaderView.ResizeMode.Interactive)
                self.table_channel_videos.setColumnWidth(i, width)
        self.table_channel_videos.setMinimumHeight(400)
        self.table_channel_videos.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table_channel_videos.customContextMenuRequested.connect(self._channel_video_table_context_menu)
        results_layout.addWidget(self.table_channel_videos)

        self.btn_export_channel_videos = QPushButton("Xuất kết quả ra Excel")
        self.btn_export_channel_videos.setFont(QFont("Arial", 10))
        self.btn_export_channel_videos.setToolTip("Xuất danh sách video của các kênh ra file Excel")
        self.btn_export_channel_videos.setFixedWidth(150)
        self.btn_export_channel_videos.clicked.connect(self._export_channel_videos_to_excel)
        self.btn_export_channel_videos.setEnabled(False)
        results_layout.addWidget(self.btn_export_channel_videos, alignment=Qt.AlignmentFlag.AlignRight)
        results_group.setLayout(results_layout)
        layout.addWidget(results_group, 1)

    def _update_channel_count(self):
        lines = self.txt_channel_urls.toPlainText().strip().splitlines()
        valid_lines = len([line for line in lines if line.strip()])
        self.lbl_channel_count.setText(f"Số kênh: {valid_lines}")

    def _start_analyze_channel_videos(self):
        if self.main_window.is_operation_running:
            QMessageBox.warning(self.main_window, "Đang xử lý", "Một tác vụ khác đang chạy. Vui lòng chờ.")
            return

        channel_urls_input = self.txt_channel_urls.toPlainText().strip().splitlines()
        if not self.main_window.api_key:
            QMessageBox.warning(self.main_window, "Lỗi API Key", "Vui lòng cấu hình API Key ở Tab 1 (API Key).")
            self.main_window.tabs.setCurrentIndex(0)
            return
        if not any(line.strip() for line in channel_urls_input):
            QMessageBox.warning(self.main_window, "Thiếu thông tin", "Vui lòng nhập ít nhất một URL hoặc ID Kênh.")
            return

        self.main_window.statusBar().showMessage("Đang phân tích các URL/ID kênh...", 0)
        QApplication.processEvents()

        self.table_channel_videos.setRowCount(0)
        self.channel_videos_data.clear()
        self.current_channel_names_for_export = []
        self.btn_export_channel_videos.setEnabled(False)

        self.main_window.is_operation_running = True
        self.main_window.update_button_states()
        self.main_window.show_progress_dialog("Đang chuẩn bị lấy video cho các kênh...", 0)

        total_channels = len([line for line in channel_urls_input if line.strip()])
        for idx, channel_url in enumerate(channel_urls_input):
            if self.main_window.current_active_thread and self.main_window.current_active_thread.isInterruptionRequested():
                break

            channel_url = channel_url.strip()
            if not channel_url:
                continue

            channel_id, error_msg = extract_channel_id_yt_dlp(
                channel_url,
                status_bar_func=self.main_window.statusBar().showMessage,
                process_events_func=QApplication.processEvents
            )
            
            if error_msg or not channel_id:
                self.main_window.statusBar().showMessage(f"Lỗi trích xuất ID kênh cho '{channel_url}': {error_msg or 'Không hợp lệ'}", 7000)
                continue

            base_progress = int(100 * idx / total_channels)
            self.main_window.update_progress_dialog(base_progress, f"Đang lấy video cho kênh {idx+1}/{total_channels}: {channel_id}...")

            self.fetch_channel_videos_thread = FetchChannelVideosThread(
                api_key=self.main_window.api_key,
                channel_id=channel_id,
                video_categories_map=self.main_window.video_categories,
                parent=self
            )
            self.fetch_channel_videos_thread.channel_videos_fetched.connect(self._on_channel_videos_fetched)
            self.fetch_channel_videos_thread.error_occurred.connect(self.main_window.on_api_error_common_slot)
            self.fetch_channel_videos_thread.progress_updated.connect(lambda p, m: self.main_window.update_progress_dialog(base_progress + int(p / total_channels), m))
            self.fetch_channel_videos_thread.finished.connect(lambda: self._on_thread_finished(total_channels, idx + 1))
            self.fetch_channel_videos_thread.start()
            self.fetch_channel_videos_thread.wait()

        self.main_window.on_worker_thread_finished()

    def _on_thread_finished(self, total_channels, current_channel_idx):
        if current_channel_idx == total_channels:
            self.main_window.hide_progress_dialog()
            self.main_window.statusBar().showMessage("Hoàn tất phân tích tất cả kênh.", 5000)
            
    def _on_channel_videos_fetched(self, videos_list, channel_name):
        if videos_list:
            self.channel_videos_data[channel_name] = videos_list
            self.current_channel_names_for_export.append(channel_name)
            self._update_display_with_filters()
            self.btn_export_channel_videos.setEnabled(True)
        else:
            self.main_window.statusBar().showMessage(f"Kênh '{channel_name}' không có video.", 3000)
    
    def _update_display_with_filters(self):
        min_views = int(self.le_min_views.text()) if self.le_min_views.text() else 0
        min_comments = int(self.le_min_comments.text()) if self.le_min_comments.text() else 0
        min_duration_minutes = int(self.le_min_duration_minutes.text()) if self.le_min_duration_minutes.text() else 0
        min_duration_seconds = min_duration_minutes * 60

        total_videos_to_process = sum(len(v) for v in self.channel_videos_data.values())
        if total_videos_to_process == 0:
            self.table_channel_videos.setRowCount(0)
            return

        self.main_window.show_progress_dialog("Đang áp dụng bộ lọc...", 0)
        
        self.table_channel_videos.setSortingEnabled(False)
        self.table_channel_videos.setRowCount(0)
        
        videos_to_display = {}
        processed_count = 0
        
        for channel_name, all_videos in self.channel_videos_data.items():
            filtered_videos = []
            for video in all_videos:
                processed_count += 1
                view_count = video.get('view_count', 0)
                comment_count_str = video.get('comment_count')
                comment_count = int(comment_count_str) if comment_count_str is not None else 0
                duration_seconds = _parse_duration_to_seconds(video.get('duration', '0'))
                
                if (view_count >= min_views and
                    comment_count >= min_comments and
                    duration_seconds >= min_duration_seconds):
                    filtered_videos.append(video)

            if filtered_videos:
                videos_to_display[channel_name] = filtered_videos

            progress_percent = int(100 * processed_count / total_videos_to_process)
            progress_message = f"Đang lọc... ({processed_count}/{total_videos_to_process})"
            self.main_window.update_progress_dialog(progress_percent, progress_message)
            QApplication.processEvents()

        for channel_name, video_list in videos_to_display.items():
            self._append_to_table(channel_name, video_list)
        
        self.table_channel_videos.setSortingEnabled(True)
        self.main_window.hide_progress_dialog()
        self.main_window.statusBar().showMessage(f"Đã lọc và hiển thị {self.table_channel_videos.rowCount()} kết quả.", 5000)

    def _append_to_table(self, channel_name, videos_list):
        start_row = self.table_channel_videos.rowCount()
        self.table_channel_videos.setRowCount(start_row + len(videos_list))

        for r_idx, video in enumerate(videos_list):
            table_row = start_row + r_idx

            channel_item = QTableWidgetItem(channel_name)
            channel_item.setFlags(channel_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table_channel_videos.setItem(table_row, 0, channel_item)

            self.table_channel_videos.setItem(table_row, 1, QTableWidgetItem(str(video.get('title', 'N/A'))))
            
            vc_val = video.get('view_count', 0)
            vc_item = QTableWidgetItem()
            vc_item.setData(Qt.ItemDataRole.DisplayRole, format_int_with_separator(vc_val))
            vc_item.setData(Qt.ItemDataRole.UserRole, vc_val)
            vc_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.table_channel_videos.setItem(table_row, 2, vc_item)

            cc_val = video.get('comment_count')
            cc_item = QTableWidgetItem()
            display_val_str = format_int_with_separator(cc_val) if cc_val is not None else "0"
            sort_val_int = int(cc_val) if cc_val is not None else 0
            
            cc_item.setData(Qt.ItemDataRole.DisplayRole, display_val_str)
            cc_item.setData(Qt.ItemDataRole.UserRole, sort_val_int)
            cc_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.table_channel_videos.setItem(table_row, 3, cc_item)
            
            upload_date = format_date_to_ddmmyyyy(video.get('upload_date', ''))
            self.table_channel_videos.setItem(table_row, 4, QTableWidgetItem(upload_date))
            
            self.table_channel_videos.setItem(table_row, 5, QTableWidgetItem(str(video.get('duration', 'N/A'))))
            self.table_channel_videos.setItem(table_row, 6, QTableWidgetItem(str(video.get('category_name', 'Không xác định'))))
            
            url = video.get('url', 'N/A')
            url_item = QTableWidgetItem(url)
            url_item.setToolTip(f"Nhấp để mở: {url}\nNhấp chuột phải để sao chép URL.")
            url_item.setForeground(QColor('blue'))
            url_item.setFont(QFont('Arial', 10, QFont.Weight.Normal, True))
            url_item.setData(Qt.ItemDataRole.UserRole, QUrl(url))
            self.table_channel_videos.setItem(table_row, 7, url_item)

            btn_open = QPushButton("Mở")
            btn_open.setFont(QFont("Arial", 9))
            btn_open.clicked.connect(lambda checked, u=video.get('url'): self.main_window.open_url_externally(u))
            self.table_channel_videos.setCellWidget(table_row, 8, btn_open)

        self.table_channel_videos.resizeRowsToContents()

    def _handle_table_cell_click(self, row, column):
        if column == 7:
            item = self.table_channel_videos.item(row, column)
            if item:
                url = item.data(Qt.ItemDataRole.UserRole)
                if url and url.isValid():
                    QDesktopServices.openUrl(url)

    def _channel_video_table_context_menu(self, pos):
        selected_item = self.table_channel_videos.itemAt(pos)
        if not selected_item: return
        column = selected_item.column()
        if column == 7:
            menu = QMenu()
            copy_url_action = menu.addAction("Sao chép URL Video")
            action = menu.exec(self.table_channel_videos.mapToGlobal(pos))
            if action == copy_url_action:
                self.main_window.copy_text_to_clipboard(selected_item.text())

    def _export_channel_videos_to_excel(self):
        if self.table_channel_videos.rowCount() == 0:
            QMessageBox.information(self.main_window, "Không có dữ liệu", "Không có dữ liệu đã lọc để xuất.")
            return

        default_filename = f"Filtered_Channels_Videos_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        
        file_path, _ = QFileDialog.getSaveFileName(
            self.main_window, "Lưu file Excel", default_filename, "Excel Files (*.xlsx)"
        )
        if not file_path:
            return

        try:
            self.main_window.statusBar().showMessage("Đang xuất kết quả đã lọc ra Excel...")
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Filtered_Results"

            # Lấy headers từ bảng (trừ cột cuối 'Hành động')
            headers = [self.table_channel_videos.horizontalHeaderItem(i).text() for i in range(self.table_channel_videos.columnCount() - 1)]
            sheet.append(headers)

            # Lặp qua từng hàng trong bảng và ghi dữ liệu ra file
            for row in range(self.table_channel_videos.rowCount()):
                row_data = []
                for col in range(len(headers)):
                    item = self.table_channel_videos.item(row, col)
                    if item:
                        # Đối với cột số, lấy dữ liệu gốc để tính toán
                        if col == 2 or col == 3: # Cột Lượt xem, Bình luận
                            value = item.data(Qt.ItemDataRole.UserRole)
                        else:
                            value = item.text()
                        row_data.append(value)
                    else:
                        row_data.append('') # Thêm ô trống nếu item không tồn tại
                sheet.append(row_data)

                # Xử lý hyperlink cho cột URL (cột thứ 8, index 7)
                url_cell = sheet.cell(row=row + 2, column=8) # +2 vì Excel row bắt đầu từ 1 và có 1 hàng header
                url = url_cell.value
                if url and "https://" in str(url):
                    url_cell.hyperlink = url
                    url_cell.font = Font(color="0000FF", underline="single")

            # Tự động điều chỉnh độ rộng cột
            for i, column_cells in enumerate(sheet.columns):
                max_length = 0
                column_letter = get_column_letter(i + 1)
                for cell in column_cells:
                    if cell.value:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                adjusted_width = min(max_length + 2, 70)
                sheet.column_dimensions[column_letter].width = adjusted_width

            workbook.save(file_path)
            self.main_window.statusBar().showMessage(f"Đã xuất thành công: {file_path}", 5000)
            QMessageBox.information(self.main_window, "Thành công",
                                   f"Dữ liệu đã lọc đã được xuất ra file:\n{file_path}")
        except Exception as e:
            self.main_window.statusBar().showMessage(f"Lỗi khi xuất Excel: {str(e)}")
            QMessageBox.critical(self.main_window, "Lỗi Xuất Excel", f"Lỗi: {str(e)}")
            traceback.print_exc()


    def set_buttons_enabled(self, enabled):
        self.btn_analyze_channel.setEnabled(enabled)
        self.btn_apply_filters.setEnabled(enabled)
        self.btn_export_channel_videos.setEnabled(enabled and self.table_channel_videos.rowCount() > 0)