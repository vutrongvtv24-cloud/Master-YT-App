import json
import logging
from datetime import datetime, timezone, timedelta
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

logger = logging.getLogger(__name__)
from isodate import parse_duration
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit,
    QPushButton, QTableWidget, QTableWidgetItem, QMessageBox,
    QFileDialog, QGroupBox, QHeaderView, QComboBox, QSpinBox,
    QAbstractSpinBox, QMenu, QCheckBox, QGridLayout, QWidgetAction
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QUrl, QSettings, QPoint
from PyQt6.QtGui import QDesktopServices, QCursor, QFont, QColor, QAction

from config import (
    YOUTUBE_REGION_LANGUAGE_MAP,
    VIDEO_DURATION_OPTIONS,
    UPLOAD_DATE_OPTIONS_DESC,
    ORDER_OPTIONS_MAP,
    VIDEO_DEFINITION_OPTIONS_MAP
)
from utils import format_date_dd_mm_yyyy, convert_iso_duration, format_int_with_separator
from services.api_manager import APIKeyManager, YouTubeService
from googleapiclient.errors import HttpError

class SearchVideosThread(QThread):
    videos_fetched = pyqtSignal(list)
    error_occurred = pyqtSignal(str)
    progress_updated = pyqtSignal(int, str)

    def __init__(self, api_key, keyword, max_results=500, region_code=None, language_code=None,
                 order='relevance', video_category_id=None,
                 video_categories_map=None, published_after_iso=None,
                 is_shorts_only=False, min_duration_seconds=0,
                 excluded_category_ids=None, parent=None):
        super().__init__(parent)
        self.api_key = api_key
        self.keyword = keyword
        self.max_results = max_results
        self.region_code = region_code
        self.language_code = language_code
        self.order = order
        self.video_category_id = video_category_id
        self.video_categories_map = video_categories_map or {}
        self.published_after_iso = published_after_iso
        self.is_shorts_only = is_shorts_only
        self.min_duration_seconds = min_duration_seconds
        self.excluded_category_ids = excluded_category_ids or []
        self._is_interruption_requested = False

    def run(self):
        if not self.api_key:
            self.error_occurred.emit("Vui lòng nhập API Key ở Tab 1 (API Key).")
            return
        if not self.keyword:
            self.error_occurred.emit("Vui lòng nhập từ khóa tìm kiếm.")
            return

        try:
            self.progress_updated.emit(0, "Đang kết nối tới YouTube...")

            # Initialize Service with Key Manager
            youtube_service_wrapper = YouTubeService()
            
            if self.isInterruptionRequested(): return
            if self.isInterruptionRequested(): return
            
            self.progress_updated.emit(10, f"Đang tìm kiếm video với từ khóa: '{self.keyword}'...")

            all_video_ids = []
            all_channel_ids = []
            next_page_token = None
            fetched_count = 0

            while fetched_count < self.max_results:
                if self.isInterruptionRequested(): return
                num_to_fetch_this_page = min(50, self.max_results - fetched_count)
                if num_to_fetch_this_page <= 0: break

                search_params = {
                    'q': self.keyword,
                    'part': 'snippet',
                    'type': 'video',
                    'maxResults': num_to_fetch_this_page,
                    'pageToken': next_page_token,
                    'order': self.order
                }
                if self.region_code: search_params['regionCode'] = self.region_code
                if self.language_code: search_params['relevanceLanguage'] = self.language_code
                if self.video_category_id: search_params['videoCategoryId'] = self.video_category_id
                if self.published_after_iso: search_params['publishedAfter'] = self.published_after_iso
                
                if self.is_shorts_only:
                    search_params['videoDuration'] = 'short'

                # Use the wrapper's search_videos which expects kwargs that match list() arguments
                # Note: 'q' is already in search_params, so we pass **search_params
                search_response = youtube_service_wrapper.search_videos(**search_params)
                if self.isInterruptionRequested(): return

                current_page_ids = []
                current_channel_ids = []
                for item in search_response.get('items', []):
                    if item.get('id', {}).get('kind') == 'youtube#video':
                        current_page_ids.append(item['id']['videoId'])
                        current_channel_ids.append(item['snippet']['channelId'])
                all_video_ids.extend(current_page_ids)
                all_channel_ids.extend(current_channel_ids)
                fetched_count += len(current_page_ids)
                next_page_token = search_response.get('nextPageToken')
                
                search_progress = 10 + int(20 * (fetched_count / self.max_results)) if self.max_results > 0 else 10
                self.progress_updated.emit(min(30, search_progress), f"Đang tìm ID video ({fetched_count}/{self.max_results})...")
                
                if not next_page_token or not current_page_ids: 
                    break
            
            if not all_video_ids:
                self.videos_fetched.emit([])
                self.progress_updated.emit(100, "Không tìm thấy video nào khớp với tiêu chí.")
                return

            self.progress_updated.emit(30, f"Đã tìm thấy {len(all_video_ids)} ID video. Đang lấy chi tiết...")

            channel_details = {}
            for i in range(0, len(all_channel_ids), 50):
                if self.isInterruptionRequested(): return
                chunk_ids = all_channel_ids[i:i+50]
                channels_response = youtube_service_wrapper.get_channel_details(
                    part='snippet,statistics',
                    id=','.join(chunk_ids)
                )
                for channel in channels_response.get('items', []):
                    stats = channel.get('statistics', {})
                    channel_details[channel['id']] = {
                        'title': channel['snippet']['title'],
                        'subscriber_count': stats.get('subscriberCount'),
                        'video_count': stats.get('videoCount'),
                        'view_count': stats.get('viewCount')
                    }

            video_details_list = []
            for i in range(0, len(all_video_ids), 50):
                if self.isInterruptionRequested(): return
                chunk_ids = all_video_ids[i:i+50]
                # Fallback to key manager for raw call if wrapper doesn't have it, 
                # OR update wrapper in parallel.
                # Let's assume we update wrapper.
                videos_response = youtube_service_wrapper.get_video_details(
                    part='snippet,statistics,contentDetails',
                    id=','.join(chunk_ids)
                )
                video_details_list.extend(videos_response.get('items', []))
                details_progress = 30 + int(60 * (len(video_details_list) / len(all_video_ids)))
                self.progress_updated.emit(details_progress, f"Đang lấy chi tiết video ({len(video_details_list)}/{len(all_video_ids)})...")

            if self.isInterruptionRequested(): return
            results = []
            for idx, video_data in enumerate(video_details_list):
                if self.isInterruptionRequested(): return
                snippet = video_data.get('snippet', {})
                statistics = video_data.get('statistics', {})
                content_details = video_data.get('contentDetails', {})
                
                duration_str = content_details.get('duration', 'PT0S')
                try:
                    duration_seconds = parse_duration(duration_str).total_seconds()
                except Exception as e:
                    logger.debug(f"Could not parse duration '{duration_str}': {e}")
                    duration_seconds = 0

                if self.is_shorts_only:
                    if duration_seconds >= 60:
                        continue
                elif self.min_duration_seconds > 0:
                    if duration_seconds < self.min_duration_seconds:
                        continue
                
                category_id = snippet.get('categoryId', '')
                if category_id in self.excluded_category_ids:
                    continue
                category_name = next((name for name, cid_val in self.video_categories_map.items() if cid_val == category_id), 'Không xác định')

                raw_comment_count = statistics.get('commentCount')
                comment_count_val = None
                if raw_comment_count is not None:
                    try:
                        comment_count_val = int(raw_comment_count)
                    except ValueError:
                        comment_count_val = None

                channel_id = snippet.get('channelId', '')
                channel_info = channel_details.get(channel_id, {
                    'title': 'N/A',
                    'subscriber_count': None,
                    'video_count': None,
                    'view_count': None
                })

                results.append({
                    'id': video_data['id'],
                    'title': snippet.get('title', 'N/A'),
                    'url': f"https://www.youtube.com/watch?v={video_data['id']}",
                    'view_count': int(statistics.get('viewCount', 0)),
                    'comment_count': comment_count_val,
                    'upload_date': snippet.get('publishedAt', 'N/A'),
                    'duration': convert_iso_duration(content_details.get('duration', 'N/A')),
                    'category_name': category_name,
                    'tags': snippet.get('tags', []),
                    'channel_title': channel_info['title'],
                    'channel_url': f"https://www.youtube.com/channel/{channel_id}" if channel_id else 'N/A',
                    'subscriber_count': channel_info['subscriber_count'],
                    'video_count': channel_info['video_count'],
                    'channel_view_count': channel_info['view_count']
                })

            self.progress_updated.emit(100, "Hoàn tất lấy thông tin video.")
            self.videos_fetched.emit(results)

        except HttpError as e:
            if self.isInterruptionRequested(): return
            try:
                error_content = json.loads(e.content.decode('utf-8'))
                error_message = error_content.get("error", {}).get("message", "Lỗi API không xác định.")
                if e.resp.status == 403 and ("quotaExceeded" in error_message or "dailyLimitExceeded" in error_message):
                    self.error_occurred.emit("Lỗi: Hạn ngạch API đã bị vượt quá.")
                else:
                    self.error_occurred.emit(f"Lỗi API: {error_message} (Code: {e.resp.status})")
            except json.JSONDecodeError:
                self.error_occurred.emit(f"Lỗi API (không thể phân tích phản hồi): {e.content.decode('utf-8', errors='ignore')}")
        except Exception as e:
            if self.isInterruptionRequested(): return
            logger.exception(f"SearchVideosThread error: {e}")
            self.error_occurred.emit(f"Lỗi không mong đợi: {str(e)}")

    def requestInterruption(self):
        self._is_interruption_requested = True
        super().requestInterruption()

    def isInterruptionRequested(self):
        return self._is_interruption_requested or super().isInterruptionRequested()

class SearchChannelsThread(QThread):
    channels_fetched = pyqtSignal(list)
    error_occurred = pyqtSignal(str)
    progress_updated = pyqtSignal(int, str)

    def __init__(self, api_key, keyword, max_results=100, region_code=None,
                 order='relevance', max_subscribers=None, max_videos=None, parent=None):
        super().__init__(parent)
        self.api_key = api_key
        self.keyword = keyword
        self.max_results = max_results
        self.region_code = region_code
        self.order = order
        self.max_subscribers = max_subscribers
        self.max_videos = max_videos
        self._is_interruption_requested = False

    def run(self):
        if not self.api_key:
            self.error_occurred.emit("Vui lòng nhập API Key ở Tab 1 (API Key).")
            return
        if not self.keyword:
            self.error_occurred.emit("Vui lòng nhập từ khóa tìm kiếm.")
            return

        try:
            self.progress_updated.emit(0, "Đang kết nối tới YouTube...")
            from googleapiclient.discovery import build
            from googleapiclient.errors import HttpError

            # Use centralized API manager (same as SearchVideosThread)
            youtube_service_wrapper = YouTubeService()

            if self.isInterruptionRequested(): return

            self.progress_updated.emit(10, f"Đang tìm kiếm kênh với từ khóa: '{self.keyword}'...")
            
            all_channel_ids = []
            next_page_token = None
            fetched_count = 0

            while fetched_count < self.max_results:
                if self.isInterruptionRequested(): return
                num_to_fetch = min(50, self.max_results - fetched_count)
                if num_to_fetch <= 0: break

                search_params = {
                    'q': self.keyword,
                    'part': 'snippet',
                    'type': 'channel',
                    'maxResults': num_to_fetch,
                    'pageToken': next_page_token,
                    'order': self.order
                }
                if self.region_code: search_params['regionCode'] = self.region_code

                search_response = youtube_service_wrapper.search_videos(self.keyword, **search_params)
                if self.isInterruptionRequested(): return

                current_page_ids = [item['id']['channelId'] for item in search_response.get('items', []) if item.get('id', {}).get('kind') == 'youtube#channel']
                all_channel_ids.extend(current_page_ids)
                
                fetched_count += len(current_page_ids)
                next_page_token = search_response.get('nextPageToken')

                progress = 10 + int(40 * (fetched_count / self.max_results))
                self.progress_updated.emit(min(50, progress), f"Đang tìm ID kênh ({fetched_count}/{self.max_results})...")

                if not next_page_token:
                    break

            if not all_channel_ids:
                self.channels_fetched.emit([])
                self.progress_updated.emit(100, "Không tìm thấy kênh nào khớp.")
                return

            self.progress_updated.emit(50, f"Đã tìm thấy {len(all_channel_ids)} kênh. Đang lấy chi tiết...")
            
            final_results = []
            for i in range(0, len(all_channel_ids), 50):
                if self.isInterruptionRequested(): return
                chunk_ids = all_channel_ids[i:i+50]
                
                channels_response = youtube_service_wrapper.get_channel_details(chunk_ids)
                
                for channel_data in channels_response.get('items', []):
                    snippet = channel_data.get('snippet', {})
                    stats = channel_data.get('statistics', {})
                    
                    subs_count = int(stats.get('subscriberCount', 0)) if stats.get('hiddenSubscriberCount') is False else -1
                    video_count = int(stats.get('videoCount', 0))
                    
                    if self.max_subscribers is not None and subs_count > self.max_subscribers and subs_count != -1:
                        continue
                    if self.max_videos is not None and video_count > self.max_videos:
                        continue

                    final_results.append({
                        'id': channel_data['id'],
                        'title': snippet.get('title', 'N/A'),
                        'description': snippet.get('description', ''),
                        'url': f"https://www.youtube.com/channel/{channel_data['id']}",
                        'subscriber_count': subs_count,
                        'video_count': video_count,
                        'view_count': int(stats.get('viewCount', 0)),
                        'published_at': snippet.get('publishedAt', 'N/A')
                    })
                
                details_progress = 50 + int(50 * ((i + len(chunk_ids)) / len(all_channel_ids)))
                self.progress_updated.emit(details_progress, f"Đang lấy chi tiết kênh ({len(final_results)})...")

            self.progress_updated.emit(100, "Hoàn tất lấy thông tin kênh.")
            self.channels_fetched.emit(final_results)

        except HttpError as e:
            if self.isInterruptionRequested(): return
            try:
                error_content = json.loads(e.content.decode('utf-8'))
                error_message = error_content.get("error", {}).get("message", "Lỗi API không xác định.")
                if e.resp.status == 403 and ("quotaExceeded" in error_message or "dailyLimitExceeded" in error_message):
                    self.error_occurred.emit("Lỗi: Hạn ngạch API đã bị vượt quá.")
                else:
                    self.error_occurred.emit(f"Lỗi API: {error_message} (Code: {e.resp.status})")
            except json.JSONDecodeError:
                self.error_occurred.emit(f"Lỗi API (không thể phân tích phản hồi): {e.content.decode('utf-8', errors='ignore')}")
        except Exception as e:
            if self.isInterruptionRequested(): return
            logger.exception(f"SearchChannelsThread error: {e}")
            self.error_occurred.emit(f"Lỗi không mong đợi: {str(e)}")

    def requestInterruption(self):
        self._is_interruption_requested = True
        super().requestInterruption()

    def isInterruptionRequested(self):
        return self._is_interruption_requested or super().isInterruptionRequested()

class KeywordResearchTab(QWidget):
    def __init__(self, main_window):
        super().__init__()
        self.main_window = main_window
        self.search_thread = None
        self.videos_data = [] 
        self.excluded_categories = []
        self.exclude_category_checkboxes = [] 
        
        # === THÊM MAP CHO BỘ LỌC NGÀY ===
        self.upload_days_map = {
            "Mặc định": 0, "1 ngày": 1, "2 ngày": 2, "7 ngày": 7, 
            "30 ngày": 30, "60 ngày": 60, "90 ngày": 90, 
            "180 ngày": 180, "365 ngày": 365,
        }
        # =================================

        self._load_excluded_categories()
        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(15, 15, 15, 15)
        layout.setSpacing(10)

        input_group = QGroupBox("Nhập liệu và Tùy chọn Tìm kiếm")
        main_input_grid = QGridLayout(input_group)
        main_input_grid.setSpacing(10)

        # Hàng 0
        main_input_grid.addWidget(QLabel("Từ khóa:"), 0, 0)
        self.txt_keyword = QLineEdit()
        self.txt_keyword.setPlaceholderText("Nhập từ khóa tìm kiếm...")
        main_input_grid.addWidget(self.txt_keyword, 0, 1, 1, 5)
        self.btn_search = QPushButton("Tìm kiếm")
        self.btn_search.clicked.connect(self._start_search)
        main_input_grid.addWidget(self.btn_search, 0, 6)

        # Hàng 1
        main_input_grid.addWidget(QLabel("Loại:"), 1, 0)
        self.combo_search_type = QComboBox()
        self.combo_search_type.addItems(["Video", "Kênh"])
        self.combo_search_type.currentIndexChanged.connect(self._toggle_filter_widgets)
        main_input_grid.addWidget(self.combo_search_type, 1, 1)
        main_input_grid.addWidget(QLabel("Quốc gia:"), 1, 2)
        self.combo_region = QComboBox()
        self.combo_region.addItems(["Tất cả"] + list(YOUTUBE_REGION_LANGUAGE_MAP.keys()))
        main_input_grid.addWidget(self.combo_region, 1, 3)
        main_input_grid.addWidget(QLabel("Sắp xếp:"), 1, 4)
        self.combo_order = QComboBox()
        self.combo_order.addItems(ORDER_OPTIONS_MAP.keys())
        main_input_grid.addWidget(self.combo_order, 1, 5, 1, 2)

        # Hàng 2
        self.video_filters_widget = QWidget()
        video_filters_layout = QHBoxLayout(self.video_filters_widget)
        video_filters_layout.setContentsMargins(0, 0, 0, 0)
        video_filters_layout.setSpacing(10)
        
        video_filters_layout.addWidget(QLabel("Danh mục:"))
        self.combo_category = QComboBox()
        video_filters_layout.addWidget(self.combo_category)

        # === THAY ĐỔI: TỪ SPINBOX SANG COMBOBOX ===
        video_filters_layout.addWidget(QLabel("Tải lên trong:"))
        self.combo_upload_days = QComboBox()
        self.combo_upload_days.addItems(self.upload_days_map.keys())
        video_filters_layout.addWidget(self.combo_upload_days)
        # =======================================
        
        video_filters_layout.addWidget(QLabel("Thời lượng tối thiểu (phút):"))
        self.spin_min_duration = QSpinBox()
        self.spin_min_duration.setRange(0, 120)
        self.spin_min_duration.setValue(0)
        self.spin_min_duration.setButtonSymbols(QAbstractSpinBox.ButtonSymbols.NoButtons)
        video_filters_layout.addWidget(self.spin_min_duration)

        self.check_shorts = QCheckBox("Chỉ tìm video Shorts")
        self.check_shorts.toggled.connect(self._toggle_duration_input)
        video_filters_layout.addWidget(self.check_shorts)

        video_filters_layout.addWidget(QLabel("Loại trừ:"))
        self.btn_exclude_categories = QPushButton("Chọn...")
        self.btn_exclude_categories.clicked.connect(self._show_exclude_category_menu)
        video_filters_layout.addWidget(self.btn_exclude_categories)
        video_filters_layout.addStretch()
        
        self.channel_filters_widget = QWidget()
        channel_filters_layout = QHBoxLayout(self.channel_filters_widget)
        channel_filters_layout.setContentsMargins(0, 0, 0, 0)
        channel_filters_layout.setSpacing(10)
        channel_filters_layout.addWidget(QLabel("Số sub tối đa:"))
        self.spin_max_subs = QSpinBox()
        self.spin_max_subs.setRange(0, 200000000); self.spin_max_subs.setValue(0)
        self.spin_max_subs.setButtonSymbols(QAbstractSpinBox.ButtonSymbols.NoButtons)
        channel_filters_layout.addWidget(self.spin_max_subs)
        channel_filters_layout.addSpacing(15)
        channel_filters_layout.addWidget(QLabel("Số video tối đa:"))
        self.spin_max_videos_channel = QSpinBox()
        self.spin_max_videos_channel.setRange(0, 1000000); self.spin_max_videos_channel.setValue(0)
        self.spin_max_videos_channel.setButtonSymbols(QAbstractSpinBox.ButtonSymbols.NoButtons)
        channel_filters_layout.addWidget(self.spin_max_videos_channel)
        channel_filters_layout.addStretch()

        main_input_grid.addWidget(self.video_filters_widget, 2, 0, 1, 7)
        main_input_grid.addWidget(self.channel_filters_widget, 2, 0, 1, 7)

        main_input_grid.setColumnStretch(1, 1)
        main_input_grid.setColumnStretch(3, 1)
        main_input_grid.setColumnStretch(5, 1)
        layout.addWidget(input_group)

        self.filter_group = QGroupBox("Bộ lọc Kết quả")
        filter_layout = QHBoxLayout(self.filter_group)
        filter_layout.addWidget(QLabel("Số view tối thiểu:"))
        self.spin_min_views = QSpinBox()
        self.spin_min_views.setRange(0, 2000000000)
        self.spin_min_views.setButtonSymbols(QAbstractSpinBox.ButtonSymbols.NoButtons)
        filter_layout.addWidget(self.spin_min_views)
        filter_layout.addSpacing(15)
        filter_layout.addWidget(QLabel("Số sub tối thiểu:"))
        self.spin_min_subs = QSpinBox()
        self.spin_min_subs.setRange(0, 2000000000)
        self.spin_min_subs.setButtonSymbols(QAbstractSpinBox.ButtonSymbols.NoButtons)
        filter_layout.addWidget(self.spin_min_subs)
        filter_layout.addSpacing(15)
        
        self.min_comments_label = QLabel("Số bình luận tối thiểu:")
        filter_layout.addWidget(self.min_comments_label)
        self.spin_min_comments = QSpinBox()
        self.spin_min_comments.setRange(0, 10000000)
        self.spin_min_comments.setButtonSymbols(QAbstractSpinBox.ButtonSymbols.NoButtons)
        filter_layout.addWidget(self.spin_min_comments)
        filter_layout.addSpacing(15)

        filter_layout.addWidget(QLabel("Trong vòng (ngày):"))
        self.spin_days_range = QSpinBox()
        self.spin_days_range.setRange(0, 10000)
        self.spin_days_range.setValue(0)
        self.spin_days_range.setButtonSymbols(QAbstractSpinBox.ButtonSymbols.NoButtons)
        filter_layout.addWidget(self.spin_days_range)
        filter_layout.addStretch()
        self.btn_apply_filter = QPushButton("Áp dụng lọc")
        self.btn_apply_filter.clicked.connect(self._apply_results_filter)
        filter_layout.addWidget(self.btn_apply_filter)
        self.btn_clear_filter = QPushButton("Xóa bộ lọc")
        self.btn_clear_filter.clicked.connect(self._clear_results_filter)
        filter_layout.addWidget(self.btn_clear_filter)
        layout.addWidget(self.filter_group)
        self.filter_group.setVisible(False)

        results_group = QGroupBox("Kết quả Tìm kiếm")
        results_layout = QVBoxLayout()
        self.table_videos = QTableWidget()
        self.table_videos.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table_videos.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table_videos.setSortingEnabled(True)
        self.table_videos.cellClicked.connect(self._handle_cell_clicked)
        self.table_videos.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table_videos.customContextMenuRequested.connect(self._video_table_context_menu)
        results_layout.addWidget(self.table_videos)
        self.btn_export_videos = QPushButton("Xuất kết quả ra Excel")
        self.btn_export_videos.clicked.connect(self._export_videos_to_excel)
        self.btn_export_videos.setEnabled(False)
        results_layout.addWidget(self.btn_export_videos, alignment=Qt.AlignmentFlag.AlignRight)
        results_group.setLayout(results_layout)
        layout.addWidget(results_group)

        self._toggle_filter_widgets()

    def _toggle_duration_input(self, checked):
        self.spin_min_duration.setEnabled(not checked)
        if checked:
            self.spin_min_duration.setValue(0)

    def _load_excluded_categories(self):
        settings = QSettings("YouTubeResearchTool", "KeywordResearchTab")
        self.excluded_categories = settings.value("excluded_categories", [], type=list)

    def _save_excluded_categories(self):
        self.excluded_categories = [cb.text() for cb in self.exclude_category_checkboxes if cb.isChecked()]
        settings = QSettings("YouTubeResearchTool", "KeywordResearchTab")
        settings.setValue("excluded_categories", self.excluded_categories)

    def update_categories_combobox(self, categories_dict):
        current_selection = self.combo_category.currentText()
        self.combo_category.clear()
        self.exclude_category_checkboxes.clear()

        if not categories_dict or not isinstance(categories_dict, dict):
            self.combo_category.addItem("Lỗi tải DM (Thử lại)")
            return

        self.combo_category.addItems(categories_dict.keys())
        
        index = self.combo_category.findText(current_selection)
        if index != -1: self.combo_category.setCurrentIndex(index)
        elif self.combo_category.count() > 0: self.combo_category.setCurrentIndex(0)

        category_names = sorted(categories_dict.keys())
        for category_name in category_names:
            checkbox = QCheckBox(category_name)
            # === SỬA LỖI 2 ===
            # ĐÃ XÓA: checkbox.setStyleSheet("color: white;")
            # Giờ đây nó sẽ sử dụng màu văn bản mặc định (thường là màu đen)
            # =================
            checkbox.setChecked(category_name in self.excluded_categories)
            self.exclude_category_checkboxes.append(checkbox)
        
        self._update_exclude_button_text()

    def _show_exclude_category_menu(self):
        if not self.exclude_category_checkboxes:
            QMessageBox.information(self, "Thông báo", "Danh sách danh mục chưa được tải. Vui lòng thử lại sau.")
            return

        menu = QMenu(self)
        
        for checkbox in self.exclude_category_checkboxes:
            action = QWidgetAction(menu)
            action.setDefaultWidget(checkbox)
            menu.addAction(action)
            checkbox.clicked.connect(lambda: menu.show())

        menu.aboutToHide.connect(self._update_exclude_button_text)
        menu.exec(self.btn_exclude_categories.mapToGlobal(QPoint(0, self.btn_exclude_categories.height())))

    def _update_exclude_button_text(self):
        self.excluded_categories = [cb.text() for cb in self.exclude_category_checkboxes if cb.isChecked()]
        count = len(self.excluded_categories)
        if count == 0:
            self.btn_exclude_categories.setText("Chọn...")
        else:
            self.btn_exclude_categories.setText(f"{count} mục")
        
        self._save_excluded_categories()

    def _toggle_filter_widgets(self):
        search_type = self.combo_search_type.currentText()
        has_data = bool(self.videos_data)
        
        self.filter_group.setVisible(has_data)
        
        if search_type == "Video":
            self.video_filters_widget.setVisible(True)
            self.channel_filters_widget.setVisible(False)
            self.min_comments_label.setVisible(True)
            self.spin_min_comments.setVisible(True)
            self._setup_video_table_headers()
        else:
            self.video_filters_widget.setVisible(False)
            self.channel_filters_widget.setVisible(True)
            self.min_comments_label.setVisible(False)
            self.spin_min_comments.setVisible(False)
            self._setup_channel_table_headers()
            
        if not has_data:
            self.table_videos.setRowCount(0)
            self.btn_export_videos.setEnabled(False)
            self.filter_group.setVisible(False)

    def _setup_video_table_headers(self):
        self.table_videos.setColumnCount(14)
        self.table_videos.setHorizontalHeaderLabels([
            "Tiêu đề", "URL Video", "Lượt xem", "Bình luận", "Ngày tải lên",
            "Thời lượng", "Tên kênh", "URL Kênh", "Số Sub của Kênh", "Số video của Kênh",
            "Số lượt Xem toàn kênh", "Danh mục", "Thẻ (Tags)", "Hành động"
        ])
        header = self.table_videos.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        col_widths = [180, 100, 100, 150, 100, 120, 150, 100, 100, 100, 100, 150, 80]
        for i, width in enumerate(col_widths, 1):
            header.setSectionResizeMode(i, QHeaderView.ResizeMode.Interactive)
            self.table_videos.setColumnWidth(i, width)
        header.setSectionResizeMode(13, QHeaderView.ResizeMode.Fixed)
        self.table_videos.setColumnWidth(13, 80)
    
    def _setup_channel_table_headers(self):
        headers = ["Tên kênh", "URL Kênh", "Số Sub", "Số Video", "Tổng lượt xem", "Ngày tạo", "Mô tả"]
        self.table_videos.setColumnCount(len(headers))
        self.table_videos.setHorizontalHeaderLabels(headers)
        header = self.table_videos.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(6, QHeaderView.ResizeMode.Stretch)
        for i in range(1, 6):
            header.setSectionResizeMode(i, QHeaderView.ResizeMode.Interactive)
        self.table_videos.resizeColumnsToContents()
    
    def _start_search(self):
        if self.main_window.is_operation_running:
            QMessageBox.warning(self.main_window, "Đang xử lý", "Một tác vụ khác đang chạy. Vui lòng chờ.")
            return

        keyword = self.txt_keyword.text().strip()
        if not self.main_window.api_key:
            QMessageBox.warning(self.main_window, "Lỗi API Key", "Vui lòng cấu hình API Key ở Tab 1 (API Key).")
            self.main_window.tabs.setCurrentIndex(0)
            return
        if not keyword:
            QMessageBox.warning(self.main_window, "Thiếu thông tin", "Vui lòng nhập từ khóa tìm kiếm.")
            return

        search_type = self.combo_search_type.currentText()
        if search_type == "Video":
            self._start_search_videos()
        else:
            self._start_search_channels()

    def _start_search_videos(self):
        region_name = self.combo_region.currentText()
        region_data = YOUTUBE_REGION_LANGUAGE_MAP.get(region_name, {})
        region_code = region_data.get("code")
        language_code = region_data.get("lang")

        max_results = 500
        order = ORDER_OPTIONS_MAP[self.combo_order.currentText()]
        category_text = self.combo_category.currentText()
        video_category_id = self.main_window.video_categories.get(category_text)
        
        # === LẤY GIÁ TRỊ TỪ COMBOBOX NGÀY ===
        selected_upload_range = self.combo_upload_days.currentText()
        upload_days = self.upload_days_map.get(selected_upload_range, 0)
        published_after_iso_str = None
        if upload_days > 0:
            published_after_dt = datetime.now(timezone.utc) - timedelta(days=upload_days)
            published_after_iso_str = published_after_dt.isoformat(timespec='seconds').replace('+00:00', 'Z')
        # ===================================
        
        is_shorts = self.check_shorts.isChecked()
        min_duration_minutes = self.spin_min_duration.value()
        min_duration_seconds = min_duration_minutes * 60

        excluded_category_ids = [self.main_window.video_categories.get(cat) for cat in self.excluded_categories if cat in self.main_window.video_categories]

        self.table_videos.setRowCount(0)
        self.videos_data = []
        self.btn_export_videos.setEnabled(False)
        self.filter_group.setVisible(False)

        self.main_window.is_operation_running = True
        self.main_window.update_button_states()
        self.main_window.show_progress_dialog(f"Đang tìm video: {self.txt_keyword.text().strip()}...", 0)

        if self.search_thread and self.search_thread.isRunning():
            self.search_thread.requestInterruption()
            self.search_thread.wait()

        self.search_thread = SearchVideosThread(
            api_key=self.main_window.api_key,
            keyword=self.txt_keyword.text().strip(),
            max_results=max_results,
            region_code=region_code,
            language_code=language_code,
            order=order,
            video_category_id=video_category_id,
            video_categories_map=self.main_window.video_categories,
            published_after_iso=published_after_iso_str,
            is_shorts_only=is_shorts,
            min_duration_seconds=min_duration_seconds,
            excluded_category_ids=excluded_category_ids,
            parent=self
        )
        self.search_thread.videos_fetched.connect(self._on_videos_fetched)
        self.search_thread.error_occurred.connect(self.main_window.on_api_error_common_slot)
        self.search_thread.progress_updated.connect(self.main_window.update_progress_dialog)
        self.search_thread.finished.connect(self.main_window.on_worker_thread_finished)
        self.search_thread.start()

    def _start_search_channels(self):
        region_name = self.combo_region.currentText()
        region_data = YOUTUBE_REGION_LANGUAGE_MAP.get(region_name, {})
        region_code = region_data.get("code")

        max_results = 500
        order = ORDER_OPTIONS_MAP[self.combo_order.currentText()]

        max_subs = self.spin_max_subs.value()
        if max_subs == 0: max_subs = None
        
        max_vids = self.spin_max_videos_channel.value()
        if max_vids == 0: max_vids = None

        self.table_videos.setRowCount(0)
        self.videos_data = []
        self.btn_export_videos.setEnabled(False)
        self.filter_group.setVisible(False) 

        self.main_window.is_operation_running = True
        self.main_window.update_button_states()
        self.main_window.show_progress_dialog(f"Đang tìm kênh: {self.txt_keyword.text().strip()}...", 0)

        if self.search_thread and self.search_thread.isRunning():
            self.search_thread.requestInterruption()
            self.search_thread.wait()

        self.search_thread = SearchChannelsThread(
            api_key=self.main_window.api_key,
            keyword=self.txt_keyword.text().strip(),
            max_results=max_results,
            region_code=region_code,
            order=order,
            max_subscribers=max_subs,
            max_videos=max_vids,
            parent=self
        )
        self.search_thread.channels_fetched.connect(self._on_channels_fetched)
        self.search_thread.error_occurred.connect(self.main_window.on_api_error_common_slot)
        self.search_thread.progress_updated.connect(self.main_window.update_progress_dialog)
        self.search_thread.finished.connect(self.main_window.on_worker_thread_finished)
        self.search_thread.start()

    def _on_videos_fetched(self, videos_list):
        self.videos_data = videos_list
        self.main_window.hide_progress_dialog()
        
        if not videos_list:
            QMessageBox.information(self.main_window, "Kết quả", "Không tìm thấy video nào khớp với tiêu chí tìm kiếm.")
            self.main_window.statusBar().showMessage("Không tìm thấy video.", 3000)
            self.btn_export_videos.setEnabled(False)
            self.filter_group.setVisible(False)
            return

        self._populate_video_table(videos_list)
        
        self.main_window.statusBar().showMessage(f"Đã tải {len(videos_list)} video.", 5000)
        QMessageBox.information(self.main_window, "Hoàn tất", f"Đã tìm thấy và hiển thị {len(videos_list)} video.")
        self.btn_export_videos.setEnabled(True)
        self.filter_group.setVisible(True)

    def _on_channels_fetched(self, channels_list):
        self.videos_data = channels_list
        self.main_window.hide_progress_dialog()

        if not channels_list:
            QMessageBox.information(self.main_window, "Kết quả", "Không tìm thấy kênh nào khớp với tiêu chí tìm kiếm.")
            self.btn_export_videos.setEnabled(False)
            self.filter_group.setVisible(False)
            return
        
        self._populate_channel_table(channels_list)

        self.main_window.statusBar().showMessage(f"Đã tải {len(channels_list)} kênh.", 5000)
        QMessageBox.information(self.main_window, "Hoàn tất", f"Đã tìm thấy và hiển thị {len(channels_list)} kênh.")
        self.btn_export_videos.setEnabled(True)
        self.filter_group.setVisible(True)

    def _populate_video_table(self, videos_list):
        self._setup_video_table_headers()
        self.table_videos.setSortingEnabled(False)
        self.table_videos.setRowCount(len(videos_list))

        for r_idx, video in enumerate(videos_list):
            self.table_videos.setItem(r_idx, 0, QTableWidgetItem(str(video.get('title', 'N/A'))))
            
            url_item = QTableWidgetItem(video.get('url', 'N/A'))
            url_item.setToolTip(f"URL: {video.get('url', 'N/A')}\nNhấp để mở hoặc chuột phải để sao chép URL.")
            url_item.setForeground(QColor('#4da6ff'))
            font = QFont(); font.setUnderline(True); url_item.setFont(font)
            self.table_videos.setItem(r_idx, 1, url_item)

            vc_val = video.get('view_count', 0)
            vc_item = QTableWidgetItem()
            vc_item.setData(Qt.ItemDataRole.DisplayRole, format_int_with_separator(vc_val))
            vc_item.setData(Qt.ItemDataRole.UserRole, vc_val)
            vc_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.table_videos.setItem(r_idx, 2, vc_item)

            cc_val = video.get('comment_count')
            cc_item = QTableWidgetItem()
            cc_display_val = format_int_with_separator(cc_val) if cc_val is not None else "N/A (Tắt)"
            cc_sort_val = cc_val if cc_val is not None else -1
            cc_item.setData(Qt.ItemDataRole.DisplayRole, cc_display_val)
            cc_item.setData(Qt.ItemDataRole.UserRole, cc_sort_val)
            cc_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.table_videos.setItem(r_idx, 3, cc_item)

            upload_date = video.get('upload_date', 'N/A')
            upload_item = QTableWidgetItem(format_date_dd_mm_yyyy(upload_date))
            if upload_date != 'N/A':
                try:
                    dt = datetime.fromisoformat(upload_date.replace('Z', '+00:00'))
                    upload_item.setData(Qt.ItemDataRole.UserRole, dt)
                except ValueError: pass
            self.table_videos.setItem(r_idx, 4, upload_item)

            self.table_videos.setItem(r_idx, 5, QTableWidgetItem(str(video.get('duration', 'N/A'))))

            channel_item = QTableWidgetItem(video.get('channel_title', 'N/A'))
            channel_item.setToolTip(f"Kênh: {video.get('channel_title', 'N/A')}\nNhấp để mở hoặc chuột phải để sao chép URL.")
            channel_item.setForeground(QColor('#4da6ff'))
            font = QFont(); font.setUnderline(True); channel_item.setFont(font)
            self.table_videos.setItem(r_idx, 6, channel_item)

            # === SỬA LỖI 3 TẠI ĐÂY ===
            channel_url_item = QTableWidgetItem(video.get('channel_url', 'N/A'))
            channel_url_item.setToolTip(f"URL Kênh: {video.get('channel_url', 'N/A')}\nNhấp để mở hoặc chuột phải để sao chép URL.")
            channel_url_item.setForeground(QColor('#4da6ff'))
            font_url = QFont(); font_url.setUnderline(True); channel_url_item.setFont(font_url)
            # Dòng bị lỗi (đã xóa): self.table_videos.setItem(r_idx, 7, channel_item)
            # Dòng chính xác:
            self.table_videos.setItem(r_idx, 7, channel_url_item)
            # ========================

            sub_val_str = video.get('subscriber_count')
            sub_val = int(sub_val_str) if sub_val_str is not None else -1
            sub_item = QTableWidgetItem()
            sub_display = format_int_with_separator(sub_val) if sub_val != -1 else "N/A"
            sub_item.setData(Qt.ItemDataRole.DisplayRole, sub_display)
            sub_item.setData(Qt.ItemDataRole.UserRole, sub_val)
            sub_item.setToolTip("Tổng số người đăng ký của kênh")
            sub_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.table_videos.setItem(r_idx, 8, sub_item)

            vid_val_str = video.get('video_count')
            vid_val = int(vid_val_str) if vid_val_str is not None else -1
            vid_item = QTableWidgetItem()
            vid_display = format_int_with_separator(vid_val) if vid_val != -1 else "N/A"
            vid_item.setData(Qt.ItemDataRole.DisplayRole, vid_display)
            vid_item.setData(Qt.ItemDataRole.UserRole, vid_val)
            vid_item.setToolTip("Tổng số video của kênh")
            vid_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.table_videos.setItem(r_idx, 9, vid_item)

            chan_view_val_str = video.get('channel_view_count')
            chan_view_val = int(chan_view_val_str) if chan_view_val_str is not None else -1
            chan_view_item = QTableWidgetItem()
            chan_view_display = format_int_with_separator(chan_view_val) if chan_view_val != -1 else "N/A"
            chan_view_item.setData(Qt.ItemDataRole.DisplayRole, chan_view_display)
            chan_view_item.setData(Qt.ItemDataRole.UserRole, chan_view_val)
            chan_view_item.setToolTip("Tổng số lượt xem của toàn kênh")
            chan_view_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.table_videos.setItem(r_idx, 10, chan_view_item)

            self.table_videos.setItem(r_idx, 11, QTableWidgetItem(str(video.get('category_name', 'Không xác định'))))
            
            tags_list = video.get('tags', [])
            tags_str = ", ".join(tags_list) if tags_list else "Không có"
            self.table_videos.setItem(r_idx, 12, QTableWidgetItem(tags_str))

            btn_open = QPushButton("Mở")
            btn_open.clicked.connect(lambda checked, u=video.get('url'): self.main_window.open_url_externally(u))
            self.table_videos.setCellWidget(r_idx, 13, btn_open)

        self.table_videos.resizeRowsToContents()
        self.table_videos.setSortingEnabled(True)

    def _populate_channel_table(self, channels_list):
        self._setup_channel_table_headers()
        self.table_videos.setSortingEnabled(False)
        self.table_videos.setRowCount(len(channels_list))

        for r_idx, channel in enumerate(channels_list):
            title_item = QTableWidgetItem(str(channel.get('title', 'N/A')))
            self.table_videos.setItem(r_idx, 0, title_item)

            url_item = QTableWidgetItem(channel.get('url', 'N/A'))
            url_item.setForeground(QColor('#4da6ff')); font = QFont(); font.setUnderline(True); url_item.setFont(font)
            self.table_videos.setItem(r_idx, 1, url_item)

            sub_val = channel.get('subscriber_count', -1)
            sub_item = QTableWidgetItem()
            sub_display = format_int_with_separator(sub_val) if sub_val != -1 else "Bị ẩn"
            sub_item.setData(Qt.ItemDataRole.DisplayRole, sub_display)
            sub_item.setData(Qt.ItemDataRole.UserRole, sub_val)
            sub_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.table_videos.setItem(r_idx, 2, sub_item)

            vid_val = channel.get('video_count', 0)
            vid_item = QTableWidgetItem()
            vid_item.setData(Qt.ItemDataRole.DisplayRole, format_int_with_separator(vid_val))
            vid_item.setData(Qt.ItemDataRole.UserRole, vid_val)
            vid_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.table_videos.setItem(r_idx, 3, vid_item)

            view_val = channel.get('view_count', 0)
            view_item = QTableWidgetItem()
            view_item.setData(Qt.ItemDataRole.DisplayRole, format_int_with_separator(view_val))
            view_item.setData(Qt.ItemDataRole.UserRole, view_val)
            view_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.table_videos.setItem(r_idx, 4, view_item)
            
            upload_date = channel.get('published_at', 'N/A')
            self.table_videos.setItem(r_idx, 5, QTableWidgetItem(format_date_dd_mm_yyyy(upload_date)))

            self.table_videos.setItem(r_idx, 6, QTableWidgetItem(str(channel.get('description', ''))))

        self.table_videos.resizeRowsToContents()
        self.table_videos.setSortingEnabled(True)

    def _apply_results_filter(self):
        if not self.videos_data:
            return

        min_views = self.spin_min_views.value()
        min_subs = self.spin_min_subs.value()
        min_comments = self.spin_min_comments.value()
        days_range = self.spin_days_range.value()

        cutoff_date = None
        if days_range > 0:
            cutoff_date = datetime.now(timezone.utc) - timedelta(days=days_range)

        filtered_data = []
        search_type = self.combo_search_type.currentText()

        for item in self.videos_data:
            passes = True
            
            if search_type == "Video":
                views = int(item.get('view_count', 0))
                subs_str = item.get('subscriber_count')
                subs = int(subs_str) if subs_str is not None else 0
                comments = item.get('comment_count') or 0
                date_str = item.get('upload_date')

                if comments < min_comments:
                    passes = False

            else: 
                views = int(item.get('view_count', 0))
                subs = int(item.get('subscriber_count', 0))
                if subs == -1: subs = 0 
                date_str = item.get('published_at') 
            
            if passes and views < min_views:
                passes = False
            if passes and subs < min_subs:
                passes = False

            if passes and cutoff_date and date_str:
                try:
                    item_date = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
                    if item_date < cutoff_date:
                        passes = False
                except (ValueError, TypeError):
                    pass 

            if passes:
                filtered_data.append(item)
        
        if search_type == "Video":
            self._populate_video_table(filtered_data)
        else:
            self._populate_channel_table(filtered_data)
            
        self.main_window.statusBar().showMessage(f"Đã áp dụng bộ lọc. Hiển thị {len(filtered_data)} kết quả.", 5000)

    def _clear_results_filter(self):
        self.spin_min_views.setValue(0)
        self.spin_min_subs.setValue(0)
        self.spin_min_comments.setValue(0)
        self.spin_days_range.setValue(0)
        
        search_type = self.combo_search_type.currentText()
        if search_type == "Video":
            self._populate_video_table(self.videos_data)
        else:
            self._populate_channel_table(self.videos_data)
        self.main_window.statusBar().showMessage("Đã xóa bộ lọc.", 3000)

    def _video_table_context_menu(self, pos):
        selected_item = self.table_videos.itemAt(pos)
        if not selected_item: return
        
        search_type = self.combo_search_type.currentText()
        column = selected_item.column()
        row_visual = selected_item.row()
        
        url_to_copy = None

        title_item = self.table_videos.item(row_visual, 0)
        if not title_item: return
        
        clicked_title = title_item.text()
        url_item = self.table_videos.item(row_visual, 1)
        if not url_item: return
        clicked_url = url_item.text()

        original_item = next((item for item in self.videos_data if item.get('url') == clicked_url and item.get('title') == clicked_title), None)
        if not original_item: return

        if search_type == "Video":
            if column in (1, 6, 7):
                url_to_copy = original_item['url'] if column == 1 else original_item['channel_url']
        else: 
            if column == 1:
                url_to_copy = original_item['url']

        if url_to_copy:
            menu = QMenu()
            copy_url_action = menu.addAction("Sao chép URL")
            action = menu.exec(self.table_videos.mapToGlobal(pos))
            if action == copy_url_action:
                self.main_window.copy_text_to_clipboard(url_to_copy)

    def _handle_cell_clicked(self, row, column):
        search_type = self.combo_search_type.currentText()
        url_to_open = None

        title_item = self.table_videos.item(row, 0)
        if not title_item: return
        
        clicked_title = title_item.text()
        url_item = self.table_videos.item(row, 1) 
        if not url_item: return
        clicked_url = url_item.text()

        original_item = next((item for item in self.videos_data if item.get('url') == clicked_url and item.get('title') == clicked_title), None)
        if not original_item: return

        if search_type == "Video":
             if column in (1, 6, 7):
                url_to_open = original_item['url'] if column == 1 else original_item['channel_url']
        else:
            if column in (0, 1):
                url_to_open = original_item['url']

        if url_to_open and url_to_open != 'N/A':
            self.main_window.open_url_externally(url_to_open)

    def _export_videos_to_excel(self):
        rows = self.table_videos.rowCount()
        cols = self.table_videos.columnCount()
        
        if rows == 0:
            QMessageBox.information(self.main_window, "Không có dữ liệu", "Không có dữ liệu để xuất.")
            return

        keyword_part = self.txt_keyword.text().strip().replace(' ', '_').replace('/', '-').replace('\\', '-')[:50]
        if not keyword_part:
            keyword_part = f"export_{datetime.now().strftime('%Y%m%d')}"
        default_filename = f"{keyword_part}.xlsx"
        
        file_path, _ = QFileDialog.getSaveFileName(
            self.main_window, "Lưu file Excel", default_filename, "Excel Files (*.xlsx)"
        )
        if not file_path:
            return

        try:
            self.main_window.statusBar().showMessage("Đang xuất ra Excel...")
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Kết quả"

            headers = [self.table_videos.horizontalHeaderItem(i).text() for i in range(cols)]
            if "Hành động" in headers:
                action_col_index = headers.index("Hành động")
                headers.pop(action_col_index)
            else:
                action_col_index = -1
            sheet.append(headers)
            
            # === SỬA LỖI 1: XUẤT SỐ RA EXCEL ===
            search_type = self.combo_search_type.currentText()
            # Chỉ số cột TRỰC QUAN (visual index) chứa dữ liệu số
            video_numeric_cols = [2, 3, 8, 9, 10]  # Lượt xem, Bình luận, Sub, Video, View kênh
            channel_numeric_cols = [2, 3, 4]       # Sub, Video, Tổng view

            for row in range(rows):
                row_data = []
                current_visual_col = 0  # Theo dõi chỉ số cột trực quan
                
                for col in range(cols):
                    if col == action_col_index:
                        continue
                        
                    item = self.table_videos.item(row, col)
                    if not item:
                        row_data.append("")
                        current_visual_col += 1
                        continue

                    is_numeric_col = False
                    if search_type == "Video" and current_visual_col in video_numeric_cols:
                        is_numeric_col = True
                    elif search_type == "Kênh" and current_visual_col in channel_numeric_cols:
                        is_numeric_col = True

                    if is_numeric_col:
                        # Lấy giá trị số thô từ UserRole
                        numeric_val = item.data(Qt.ItemDataRole.UserRole)
                        # Giá trị -1 được dùng cho "N/A" hoặc "Bị ẩn"
                        if numeric_val is not None and numeric_val != -1:
                            row_data.append(numeric_val)
                        else:
                            # Giữ nguyên văn bản "N/A", "Bị ẩn" v.v.
                            row_data.append(item.text())
                    else:
                        # Lấy giá trị văn bản như bình thường
                        row_data.append(item.text())
                        
                    current_visual_col += 1
                sheet.append(row_data)
            # =================================

            for i, column_cells in enumerate(sheet.columns):
                max_length = 0
                column_letter = get_column_letter(i + 1)
                for cell in column_cells:
                    if cell.value:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except Exception:
                            pass  # Non-critical: ignore errors when calculating column width
                adjusted_width = min(max_length + 2, 70)
                sheet.column_dimensions[column_letter].width = adjusted_width

            workbook.save(file_path)
            self.main_window.statusBar().showMessage(f"Đã xuất thành công ra: {file_path}", 5000)
            QMessageBox.information(self.main_window, "Thành công", f"Dữ liệu đã được xuất ra file:\n{file_path}")

        except Exception as e:
            self.main_window.statusBar().showMessage(f"Lỗi khi xuất Excel: {str(e)}")
            QMessageBox.critical(self.main_window, "Lỗi Xuất Excel", f"Lỗi: {str(e)}")
            logger.exception(f"Export videos to Excel error: {e}")

    def set_buttons_enabled(self, enabled):
        self.btn_search.setEnabled(enabled)
        has_data = self.table_videos.rowCount() > 0
        self.btn_export_videos.setEnabled(enabled and has_data)