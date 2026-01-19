# youtube_research_tool/ui_tabs/tab_suggestions.py

import json
import traceback
import requests # For YouTube suggestions API (non-data API)
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit,
    QPushButton, QMessageBox, QFileDialog, QGroupBox,
    QListWidget, QListWidgetItem,QComboBox
)
from PyQt6.QtCore import QThread, pyqtSignal, Qt

# Import từ các tệp trong dự án
from config import YOUTUBE_REGION_LANGUAGE_MAP

class FetchSuggestionsThread(QThread):
    suggestions_fetched = pyqtSignal(list)
    error_occurred = pyqtSignal(str)
    progress_updated = pyqtSignal(int, str) # int for percentage, str for message

    def __init__(self, seed_keyword, gl_code=None, hl_code=None, parent=None):
        super().__init__(parent)
        self.seed_keyword = seed_keyword
        self.gl_code = gl_code
        self.hl_code = hl_code
        self._is_interruption_requested = False

    def run(self):
        if not self.seed_keyword:
            self.error_occurred.emit("Từ khóa gốc không được để trống.")
            return

        if self.isInterruptionRequested():
            self.error_occurred.emit("Thao tác lấy gợi ý đã bị hủy.") # Or just return silently
            return

        region_info_parts = []
        if self.gl_code:
            region_info_parts.append(f"Vùng: {self.gl_code}")
        if self.hl_code:
            region_info_parts.append(f"Ngôn ngữ: {self.hl_code}")
        
        region_info_str = ""
        if region_info_parts:
            region_info_str = f" ({', '.join(region_info_parts)})"

        self.progress_updated.emit(0, f"Đang lấy gợi ý cho: '{self.seed_keyword}'{region_info_str}...")

        try:
            params = {"client": "firefox", "ds": "yt", "q": self.seed_keyword}
            if self.gl_code:
                params["gl"] = self.gl_code
            if self.hl_code:
                params["hl"] = self.hl_code

            if self.isInterruptionRequested(): return

            # Use a session for potentially better connection handling
            session = requests.Session()
            response = session.get(
                "http://suggestqueries.google.com/complete/search",
                params=params,
                timeout=10  # seconds
            )
            response.raise_for_status()  # Raises an HTTPError for bad responses (4XX or 5XX)

            if self.isInterruptionRequested(): return

            data = response.json()

            # Expected format: ["query", ["suggestion1", "suggestion2", ...], [descriptions...], {metadata...}]
            # We are interested in data[1]
            if isinstance(data, list) and len(data) >= 2 and isinstance(data[1], list):
                suggestions = [str(s) for s in data[1]] # Ensure all are strings
                self.progress_updated.emit(100, f"Đã tìm thấy {len(suggestions)} gợi ý.")
                self.suggestions_fetched.emit(suggestions)
            elif isinstance(data, list) and data and isinstance(data[0], str) and \
                 (len(data) == 1 or not data[1]): # Case where original query is returned but no other suggestions
                self.progress_updated.emit(100, "Không tìm thấy gợi ý bổ sung (chỉ có từ khóa gốc).")
                self.suggestions_fetched.emit([]) # No additional suggestions
            else:
                self.error_occurred.emit(f"Định dạng phản hồi gợi ý không mong đợi. Dữ liệu: {str(data)[:200]}")

        except requests.exceptions.Timeout:
            if self.isInterruptionRequested(): return
            self.error_occurred.emit("Lỗi: Hết thời gian chờ khi lấy gợi ý.")
        except requests.exceptions.RequestException as e:
            if self.isInterruptionRequested(): return
            self.error_occurred.emit(f"Lỗi mạng hoặc HTTP khi lấy gợi ý: {str(e)}")
        except json.JSONDecodeError:
            if self.isInterruptionRequested(): return
            # Try to show part of the response if it's not JSON
            raw_response = "Không thể đọc"
            if 'response' in locals() and hasattr(response, 'text'):
                raw_response = response.text[:200] # Get first 200 chars
            self.error_occurred.emit(f"Lỗi: Không thể phân tích phản hồi JSON từ máy chủ gợi ý. Phản hồi nhận được (một phần): {raw_response}")
        except Exception as e:
            if self.isInterruptionRequested(): return
            traceback.print_exc()
            self.error_occurred.emit(f"Lỗi không mong đợi khi lấy gợi ý: {str(e)}.")

    def requestInterruption(self):
        self._is_interruption_requested = True
        super().requestInterruption()

    def isInterruptionRequested(self):
        return self._is_interruption_requested or super().isInterruptionRequested()


class SuggestionsTab(QWidget):
    def __init__(self, main_window):
        super().__init__()
        self.main_window = main_window
        self.fetch_suggestions_thread = None
        self.suggested_keywords_data = [] # Store fetched suggestions for this tab

        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(15, 15, 15, 15)
        layout.setSpacing(10)

        # --- Input Group ---
        input_group = QGroupBox("Nhập từ khóa và Tùy chọn gợi ý")
        input_main_layout = QVBoxLayout()

        seed_layout = QHBoxLayout()
        seed_layout.addWidget(QLabel("Từ khóa gốc:"))
        self.txt_seed_keyword = QLineEdit()
        self.txt_seed_keyword.setPlaceholderText("Nhập từ khóa để lấy gợi ý...")
        seed_layout.addWidget(self.txt_seed_keyword, 1)
        input_main_layout.addLayout(seed_layout)

        sug_region_layout = QHBoxLayout()
        sug_region_layout.addWidget(QLabel("Quốc gia & Ngôn ngữ (cho gợi ý):"))
        self.combo_suggestions_region = QComboBox()
        self.combo_suggestions_region.addItems(YOUTUBE_REGION_LANGUAGE_MAP.keys())
        sug_region_layout.addWidget(self.combo_suggestions_region, 1)
        input_main_layout.addLayout(sug_region_layout)

        fetch_btn_layout = QHBoxLayout()
        fetch_btn_layout.addStretch()
        self.btn_fetch_suggestions = QPushButton("Lấy gợi ý")
        self.btn_fetch_suggestions.clicked.connect(self._start_fetch_suggestions)
        fetch_btn_layout.addWidget(self.btn_fetch_suggestions)
        input_main_layout.addLayout(fetch_btn_layout)

        input_group.setLayout(input_main_layout)
        layout.addWidget(input_group)

        # --- Results Group ---
        results_group = QGroupBox("Từ khóa gợi ý")
        results_layout = QVBoxLayout()
        self.list_suggestions = QListWidget()
        self.list_suggestions.setToolTip("Danh sách các từ khóa gợi ý từ YouTube.")
        results_layout.addWidget(self.list_suggestions)

        self.btn_export_suggestions = QPushButton("Xuất gợi ý ra Excel")
        self.btn_export_suggestions.clicked.connect(self._export_suggestions_to_excel)
        self.btn_export_suggestions.setEnabled(False)
        results_layout.addWidget(self.btn_export_suggestions, alignment=Qt.AlignmentFlag.AlignRight)
        results_group.setLayout(results_layout)
        layout.addWidget(results_group)
        layout.addStretch() # Ensure the results group doesn't take all vertical space if list is short

    def _start_fetch_suggestions(self):
        if self.main_window.is_operation_running:
            QMessageBox.warning(self.main_window, "Đang xử lý", "Một tác vụ khác đang chạy. Vui lòng chờ.")
            return

        seed_keyword = self.txt_seed_keyword.text().strip()
        if not seed_keyword:
            QMessageBox.warning(self.main_window, "Thiếu thông tin", "Vui lòng nhập từ khóa gốc.")
            return

        sug_region_name = self.combo_suggestions_region.currentText()
        sug_region_data = YOUTUBE_REGION_LANGUAGE_MAP.get(sug_region_name, {})
        gl_code = sug_region_data.get("code")
        hl_code = sug_region_data.get("lang")

        self.list_suggestions.clear()
        self.suggested_keywords_data = [] # Clear previous data
        self.btn_export_suggestions.setEnabled(False)

        self.main_window.is_operation_running = True
        self.main_window.update_button_states()
        
        status_parts = [f"Đang lấy gợi ý cho '{seed_keyword}'"]
        if gl_code: status_parts.append(f"Vùng: {gl_code}")
        if hl_code: status_parts.append(f"Ngôn ngữ: {hl_code}")
        status_msg = ", ".join(status_parts) + "..."
        self.main_window.show_progress_dialog(status_msg, 0) # Show progress dialog

        if self.fetch_suggestions_thread and self.fetch_suggestions_thread.isRunning():
            self.fetch_suggestions_thread.requestInterruption()
            self.fetch_suggestions_thread.wait()

        self.fetch_suggestions_thread = FetchSuggestionsThread(
            seed_keyword=seed_keyword,
            gl_code=gl_code,
            hl_code=hl_code,
            parent=self
        )
        self.fetch_suggestions_thread.suggestions_fetched.connect(self._on_suggestions_fetched)
        self.fetch_suggestions_thread.error_occurred.connect(self.main_window.on_api_error_common_slot)
        self.fetch_suggestions_thread.progress_updated.connect(self.main_window.update_progress_dialog)
        self.fetch_suggestions_thread.finished.connect(self.main_window.on_worker_thread_finished)
        self.fetch_suggestions_thread.start()

    def _on_suggestions_fetched(self, suggestions_list):
        self.suggested_keywords_data = suggestions_list
        self.main_window.hide_progress_dialog()

        if not suggestions_list:
            self.main_window.statusBar().showMessage("Không tìm thấy gợi ý nào.", 3000)
            QMessageBox.information(self.main_window, "Thông báo", "Không tìm thấy gợi ý nào cho từ khóa này.")
            self.btn_export_suggestions.setEnabled(False)
            return

        self.list_suggestions.clear()
        for item_text in suggestions_list:
            self.list_suggestions.addItem(QListWidgetItem(item_text))
            
        self.main_window.statusBar().showMessage(f"Đã tải {len(suggestions_list)} từ khóa gợi ý.", 3000)
        QMessageBox.information(self.main_window, "Hoàn tất", f"Đã tìm thấy và hiển thị {len(suggestions_list)} gợi ý.")
        self.btn_export_suggestions.setEnabled(True)

    def _export_suggestions_to_excel(self):
        if not self.suggested_keywords_data:
            QMessageBox.information(self.main_window, "Không có dữ liệu", "Không có gợi ý để xuất.")
            return

        seed_part = self.txt_seed_keyword.text().strip().replace(' ', '_').replace('/', '-').replace('\\','-')[:30]
        default_filename = f"Youtube_Suggestions_{seed_part if seed_part else 'data'}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        
        file_path, _ = QFileDialog.getSaveFileName(
            self.main_window, "Lưu file Excel Gợi ý", default_filename, "Excel Files (*.xlsx)"
        )
        if not file_path:
            return

        try:
            self.main_window.statusBar().showMessage("Đang xuất gợi ý ra Excel...")
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Từ khóa Gợi ý"
            sheet.append(["Từ khóa gợi ý"]) # Header

            for keyword in self.suggested_keywords_data:
                sheet.append([keyword])
            
            # Adjust column width for the first column
            if sheet.max_column >= 1:
                max_length = 0
                column_letter = get_column_letter(1) # Column 'A'
                for cell in sheet[column_letter]:
                    if cell.value:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                adjusted_width = min(max_length + 5, 100) # Add some padding, max width 100
                sheet.column_dimensions[column_letter].width = adjusted_width

            workbook.save(file_path)
            self.main_window.statusBar().showMessage(f"Đã xuất gợi ý thành công: {file_path}", 5000)
            QMessageBox.information(self.main_window, "Thành công",
                                   f"Dữ liệu gợi ý ({len(self.suggested_keywords_data)}) đã được xuất ra file:\n{file_path}")
        except Exception as e:
            self.main_window.statusBar().showMessage(f"Lỗi khi xuất gợi ý Excel: {str(e)}")
            QMessageBox.critical(self.main_window, "Lỗi Xuất Excel", f"Lỗi: {str(e)}")
            traceback.print_exc()

    def set_buttons_enabled(self, enabled):
        self.btn_fetch_suggestions.setEnabled(enabled)
        self.btn_export_suggestions.setEnabled(enabled and bool(self.suggested_keywords_data))